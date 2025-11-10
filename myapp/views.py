from io import BytesIO
from zipfile import ZipFile
import json
import random
import io
from django.shortcuts import redirect
import string
from django.db.models import Sum
from django.shortcuts import render, redirect
from django.urls import reverse
from DSI2025 import settings
from .forms import *
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from functools import wraps
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from .models import Pelicula, Reserva, Valoracion,CodigoDescuento
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.db.models import Q
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, Table, TableStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from io import BytesIO
from django.conf import settings
import os
from datetime import datetime
import qrcode
from reportlab.platypus import Image as RLImage
from PIL import Image as PILImage
from datetime import date
from django.db import models
from itertools import groupby
from django.shortcuts import render
from django.urls import reverse
from django.db.models import Q

from datetime import date
from django.contrib import messages
from django.db import transaction
from django.shortcuts import get_object_or_404, render, redirect
from django.views.decorators.csrf import csrf_exempt
from .models import Pelicula, Funcion, Reserva, User
from .decorators import admin_required
from django.http import HttpResponse
import pandas as pd
from reportlab.pdfgen import canvas
from .models import Pelicula, Venta
from django.db.models import Sum
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from .models import Venta
from django.db.models import Prefetch
from datetime import datetime
from .email import send_brevo_email   
from django.utils import timezone 
from datetime import date, datetime, timedelta
from itertools import groupby
import pytz
from decimal import Decimal

from .models import Pelicula, Funcion, Pago, MetodoPago, Valoracion
from .utils.payment_simulator import simular_pago
from .utils.encryption import encrypt_card_data, encrypt_card_data_full, decrypt_card_data, get_card_type
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Avg, Count
from reportlab.lib.pagesizes import A4, letter
import openpyxl 
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import date

# Diccionario de géneros con nombres completos
GENERO_CHOICES_DICT = {
    "AC": "Acción",
    "DR": "Drama",
    "CO": "Comedia",
    "TE": "Terror",
    "CF": "Ciencia Ficción",
    "RO": "Romance",
    "DO": "Documental",
    "AN": "Animacion",
    "FA": "Fantasía",  
}

@admin_required
@staff_member_required
@csrf_exempt
def peliculas(request):
    from datetime import date
    hoy = date.today()

    # 🔹 Filtrar películas según fecha de estreno
    peliculas_en_cartelera = Pelicula.objects.filter(
        Q(fecha_estreno__lte=hoy) | Q(fecha_estreno__isnull=True)
    ).order_by('-id')

    peliculas_proximas = Pelicula.objects.filter(
        fecha_estreno__gt=hoy
    ).order_by('fecha_estreno')

    # 🔹 Procesar búsqueda si existe
    busqueda = request.GET.get('busqueda', '').strip()
    if busqueda:
        peliculas_en_cartelera = peliculas_en_cartelera.filter(
            Q(nombre__icontains=busqueda) | Q(director__icontains=busqueda)
        )
        peliculas_proximas = peliculas_proximas.filter(
            Q(nombre__icontains=busqueda) | Q(director__icontains=busqueda)
        )

    # 🔹 Procesar formulario (crear / editar / eliminar)
    if request.method == 'POST':
        accion = request.POST.get('accion')

        # --- CREAR ---
        if accion == 'crear':
            nombre = request.POST.get('nombre', '').strip()
            anio = request.POST.get('anio', '').strip()
            director = request.POST.get('director', '').strip()
            imagen_url = request.POST.get('imagen_url', '').strip()
            trailer_url = request.POST.get('trailer_url', '').strip()
            generos = request.POST.getlist('generos')
            salas = request.POST.getlist('salas')
            fecha_estreno = request.POST.get('fecha_estreno', '').strip()
            clasificacion = request.POST.get('clasificacion', 'APT')
            idioma = request.POST.get('idioma', 'ESP')

            errores = []
            if not nombre:
                errores.append('El nombre es obligatorio.')
            if Pelicula.objects.filter(nombre=nombre).exists():
                errores.append('Ya existe una película con ese nombre.')
            if not anio.isdigit() or int(anio) < 1900 or int(anio) > 2099:
                errores.append('El año debe estar entre 1900 y 2099.')
            if not director:
                errores.append('El director es obligatorio.')
            if not generos:
                errores.append('Debe seleccionar al menos un género.')
            if len(generos) > 3:
                errores.append('No puede seleccionar más de 3 géneros.')

            if errores:
                for e in errores:
                    messages.error(request, e)
            else:
                pelicula = Pelicula(
                    nombre=nombre,
                    anio=anio,
                    director=director,
                    imagen_url=imagen_url,
                    trailer_url=trailer_url,
                    generos=",".join(generos),
                    salas=",".join(salas),
                    fecha_estreno=fecha_estreno if fecha_estreno else None,
                    clasificacion=clasificacion,
                    idioma=idioma
                )
                pelicula.save()
                messages.success(request, f'Película "{nombre}" creada exitosamente.')
                return redirect('peliculas')

        # --- EDITAR ---
        elif accion == 'editar':
            pelicula_id = request.POST.get('pelicula_id', '').strip()
            
            if not pelicula_id:
                messages.error(request, 'ID de película no proporcionado.')
                return redirect('peliculas')
            
            try:
                pelicula = Pelicula.objects.get(id=pelicula_id)
            except Pelicula.DoesNotExist:
                messages.error(request, 'No se encontró la película a editar.')
                return redirect('peliculas')

            # Validar nombre duplicado (excepto la misma película)
            nuevo_nombre = request.POST.get('nombre', '').strip()
            if Pelicula.objects.filter(nombre=nuevo_nombre).exclude(id=pelicula.id).exists():
                messages.error(request, 'Ya existe otra película con ese nombre.')
                return redirect('peliculas')

            pelicula.nombre = nuevo_nombre
            pelicula.anio = request.POST.get('anio', '').strip()
            pelicula.director = request.POST.get('director', '').strip()
            pelicula.imagen_url = request.POST.get('imagen_url', '').strip()
            pelicula.trailer_url = request.POST.get('trailer_url', '').strip()
            pelicula.generos = ",".join(request.POST.getlist('generos'))
            pelicula.salas = ",".join(request.POST.getlist('salas'))
            pelicula.fecha_estreno = request.POST.get('fecha_estreno') or None
            pelicula.clasificacion = request.POST.get('clasificacion', 'APT')
            pelicula.idioma = request.POST.get('idioma', 'ESP')
            pelicula.save()
            messages.success(request, f'Película "{pelicula.nombre}" actualizada correctamente.')
            return redirect('peliculas')

        # --- ELIMINAR ---
        elif accion == 'eliminar':
            pelicula_id = request.POST.get('pelicula_id', '').strip()
            
            if not pelicula_id:
                messages.error(request, 'ID de película no proporcionado.')
                return redirect('peliculas')
            
            try:
                pelicula = Pelicula.objects.get(id=pelicula_id)
                nombre_pelicula = pelicula.nombre
                pelicula.delete()
                messages.success(request, f'Película "{nombre_pelicula}" eliminada correctamente.')
            except Pelicula.DoesNotExist:
                messages.error(request, 'No se encontró la película para eliminar.')
            return redirect('peliculas')

    # 🔹 Datos para el formulario y la tabla
    generos_choices = dict(Pelicula.GENERO_CHOICES)
    salas_disponibles = Pelicula.SALAS_DISPONIBLES

    pelicula_editar = None
    if 'editar' in request.GET:
        pelicula_id = request.GET.get('editar')
        if pelicula_id:
            pelicula_editar = Pelicula.objects.filter(id=pelicula_id).first()

    # 🔹 Convertir películas de cartelera con sus salas y formatos
    peliculas_con_pares = []
    for p in peliculas_en_cartelera:
        pares = p.get_salas_con_formato()
        generos_nombres = [generos_choices.get(g, g) for g in p.get_generos_list()]
        peliculas_con_pares.append({
            'obj': p,
            'pares': pares,
            'generos_nombres': ", ".join(generos_nombres),
            'clasificacion': p.clasificacion,
            'idioma': p.idioma
        })

    # 🔹 Contexto para renderizar
    context = {
        'peliculas': peliculas_con_pares,
        'peliculas_proximas': peliculas_proximas,
        'GENERO_CHOICES_DICT': generos_choices,
        'SALAS_DISPONIBLES': salas_disponibles,
        'pelicula_editar': pelicula_editar,
        'busqueda': busqueda,
    }

    return render(request, 'peliculas.html', context)

#####################################################################

# Decorador personalizado para verificar si el usuario es admin
def admin_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            # Redirigir a nuestro login personalizado
            return redirect('/accounts/login/')
        if not (request.user.is_staff or request.user.is_superuser):
            # Si no es superuser, redirigir al índice
            return redirect('index')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

#####################################################################################
def convertir_generos(codigos_generos):
    """Convierte códigos de género a nombres completos"""
    if not codigos_generos:
        return []
    return [GENERO_CHOICES_DICT.get(codigo.strip(), "Desconocido") 
            for codigo in codigos_generos.split(",")]

###############################################################################3

def index(request):
    from itertools import groupby
    from django.http import JsonResponse
    from django.db import models
    
    # ✅ Usar datetime.now() sin zona horaria
    ahora_naive = datetime.now()
    hoy = ahora_naive.date()
    
    # ✅ Obtener fecha seleccionada (si existe)
    fecha_seleccionada_str = request.GET.get('fecha')
    if fecha_seleccionada_str:
        try:
            fecha_seleccionada = datetime.strptime(fecha_seleccionada_str, '%Y-%m-%d').date()
        except ValueError:
            fecha_seleccionada = hoy
    else:
        fecha_seleccionada = hoy

    # ✅ Nombres de meses en español
    nombres_meses_es = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto', 
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    
    nombres_dias_es = {
        'Monday': 'Lunes',
        'Tuesday': 'Martes', 
        'Wednesday': 'Miércoles',
        'Thursday': 'Jueves',
        'Friday': 'Viernes',
        'Saturday': 'Sábado',
        'Sunday': 'Domingo'
    }
    
    # ✅ Generar lista de 5 días (hoy + 4 días siguientes)
    dias_disponibles = []
    for i in range(5):
        dia = hoy + timedelta(days=i)
        dia_info = {
            'fecha': dia,
            'nombre_es': nombres_dias_es[dia.strftime('%A')],
            'formato_corto': dia.strftime('%d/%m')
        }
        dias_disponibles.append(dia_info)

    # ✅ Información de fecha seleccionada en español (FORMATO COMPLETO)
    nombre_dia_seleccionado = nombres_dias_es[fecha_seleccionada.strftime('%A')]
    nombre_mes_seleccionado = nombres_meses_es[fecha_seleccionada.month]
    
    # Formato: "Lunes 03 de Noviembre"
    fecha_formateada = f"{nombre_dia_seleccionado} {fecha_seleccionada.day:02d} de {nombre_mes_seleccionado}"

    # 🔴 CORRECCIÓN CRÍTICA: Filtrar funciones que estén vigentes en la fecha seleccionada
    # Una función está vigente si:
    # 1. está activa (activa=True)
    # 2. ya comenzó (fecha_inicio <= fecha_seleccionada)
    # 3. NO ha terminado (fecha_fin >= fecha_seleccionada)
    
    from django.db.models import F, ExpressionWrapper, DurationField
    
    
    funciones = (
        Funcion.objects.filter(
            activa=True,
            fecha_inicio__lte=fecha_seleccionada,
        )
        .select_related('pelicula')
        .order_by('-pelicula__id', 'horario')
    )

    # 🔴 Filtrar funciones que NO hayan expirado para la fecha seleccionada
    funciones_filtradas = []
    for funcion in funciones:
        # Calcular fecha de fin: fecha_inicio + semanas - 1 día
        fecha_fin_funcion = funcion.fecha_inicio + timedelta(weeks=funcion.semanas) - timedelta(days=1)
        
        # ✅ CORRECCIÓN: Verificar que la función NO haya terminado
        if fecha_fin_funcion < fecha_seleccionada:
            continue  # Esta función ya terminó, no la incluimos
        
        # Si es hoy, verificar que el horario no haya pasado (con margen de 10 min)
        if fecha_seleccionada == hoy:
            hora_funcion = datetime.strptime(funcion.horario, '%H:%M').time()
            datetime_funcion = datetime.combine(fecha_seleccionada, hora_funcion)
            
            margen = timedelta(minutes=10)
            if datetime_funcion > (ahora_naive - margen):
                funciones_filtradas.append(funcion)
        else:
            # Para fechas futuras, incluir todas las funciones vigentes
            funciones_filtradas.append(funcion)

    # ✅ Agrupar funciones por película (ordenadas por ID descendente)
    peliculas_cartelera = []
    peliculas_vistas = set()
    
    for funcion in funciones_filtradas:
        if funcion.pelicula.id not in peliculas_vistas:
            # Primera vez que vemos esta película
            pelicula = funcion.pelicula
            pelicula.funciones_del_dia = [f for f in funciones_filtradas if f.pelicula.id == pelicula.id]
            peliculas_cartelera.append(pelicula)
            peliculas_vistas.add(pelicula.id)

    # ✅ Enriquecer películas de cartelera
    for pelicula in peliculas_cartelera:
        pelicula.generos_list = pelicula.get_generos_list()
        pelicula.salas_con_formato = pelicula.get_salas_con_formato()
        pelicula.estrellas = (
            pelicula.get_rating_estrellas()
            if hasattr(pelicula, 'get_rating_estrellas')
            else {'llenas': 0, 'media': False}
        )

    # ✅ Si es request AJAX, devolver JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        peliculas_data = []
        for pelicula in peliculas_cartelera:
            # Obtener imagen URL
            imagen_url = ''
            if hasattr(pelicula, 'poster') and pelicula.poster:
                imagen_url = pelicula.poster.url
            elif hasattr(pelicula, 'imagen_url'):
                imagen_url = pelicula.imagen_url
            
            peliculas_data.append({
                'id': pelicula.id,
                'nombre': pelicula.nombre,
                'imagen_url': imagen_url,
                'generos': pelicula.generos_list if hasattr(pelicula, 'generos_list') else [],
                'director': pelicula.director if hasattr(pelicula, 'director') else 'Desconocido',
                'clasificacion': pelicula.get_clasificacion_display() if hasattr(pelicula, 'get_clasificacion_display') else 'N/A',
                'idioma': pelicula.get_idioma_display() if hasattr(pelicula, 'get_idioma_display') else 'N/A',
                'anio': pelicula.anio if hasattr(pelicula, 'anio') else '',
                'trailer_url': pelicula.trailer_url if hasattr(pelicula, 'trailer_url') else '#',
                'rating': pelicula.estrellas if hasattr(pelicula, 'estrellas') else {'llenas': 0, 'media': False},
                'rating_promedio': str(pelicula.get_rating_promedio()) if hasattr(pelicula, 'get_rating_promedio') else '0.0',
                'total_valoraciones': pelicula.get_total_valoraciones() if hasattr(pelicula, 'get_total_valoraciones') else 0,
                'funciones': [{
                    'horario': funcion.get_horario_display() if hasattr(funcion, 'get_horario_display') else funcion.horario,
                    'sala': str(funcion.sala),
                    'formato': funcion.get_formato_sala() if hasattr(funcion, 'get_formato_sala') else ''
                } for funcion in pelicula.funciones_del_dia]
            })
        
        return JsonResponse({
            'fecha_formateada': fecha_formateada,
            'es_hoy': fecha_seleccionada == hoy,
            'peliculas': peliculas_data,
            'total_peliculas': len(peliculas_cartelera)
        })

    # ✅ BASE DE DATOS (Solo Admin): TODAS las películas en cartelera
    peliculas_base_datos = []
    if request.user.is_authenticated and request.user.is_staff:
        peliculas_base_datos = Pelicula.objects.filter(
            models.Q(fecha_estreno__lte=hoy) | models.Q(fecha_estreno__isnull=True)
        ).order_by('-id')  # Más recientes primero
        
        # Enriquecer todas las películas de base de datos
        for pelicula in peliculas_base_datos:
            pelicula.generos_list = pelicula.get_generos_list()
            pelicula.salas_con_formato = pelicula.get_salas_con_formato()
            pelicula.estrellas = (
                pelicula.get_rating_estrellas()
                if hasattr(pelicula, 'get_rating_estrellas')
                else {'llenas': 0, 'media': False}
            )
            # Verificar si tiene funciones activas para la fecha seleccionada
            pelicula.tiene_funciones = Funcion.objects.filter(
                pelicula=pelicula,
                activa=True,
                fecha_inicio__lte=fecha_seleccionada
            ).exists()

    # ✅ PRÓXIMAMENTE: Películas con fecha de estreno futura
    peliculas_proximas = Pelicula.objects.filter(
        fecha_estreno__gt=hoy
    ).order_by('fecha_estreno')

    # Enriquecer películas próximas
    for pelicula in peliculas_proximas:
        pelicula.generos_list = pelicula.get_generos_list()

    es_admin = request.user.is_authenticated and request.user.is_staff

    return render(request, 'index.html', {
        'peliculas_base_datos': peliculas_base_datos,
        'peliculas_proximas': peliculas_proximas,
        'peliculas_cartelera': peliculas_cartelera,
        'es_admin': es_admin,
        'dias_disponibles': dias_disponibles,
        'fecha_seleccionada': fecha_seleccionada,
        'fecha_formateada': fecha_formateada,
        'hoy': hoy,
        'nombres_dias_es': nombres_dias_es,
        'nombres_meses_es': nombres_meses_es,
    })

################################################################################



@staff_member_required
def panel_admin(request):
    """Vista del panel de administración"""
    return render(request, 'panelAdmin.html')


##################################################################

def my_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        remember_me = request.POST.get('remember_me')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            
            if remember_me:
                request.session.set_expiry(settings.SESSION_COOKIE_AGE)
            else:
                request.session.set_expiry(0)
            
            
            # Redirigir según el tipo de usuario
            if user.is_superuser or user.is_staff:
                return redirect('panelAdmin')  # ✅ Sin barra
            else:
                return redirect('index')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
    
    return render(request, 'registration/login.html')


def my_logout(request):
    """Vista para cerrar sesión del usuario"""
    from django.contrib.auth import logout
    
    if request.method == 'POST':
        logout(request)
        messages.success(request, 'Has cerrado sesión exitosamente.')
        return redirect('index')
    elif request.method == 'GET':
        # Para enlaces directos, también permitimos logout via GET
        logout(request)
        messages.success(request, 'Has cerrado sesión exitosamente.')
        return redirect('index')
    
    return redirect('index')


def registro_usuario(request):
    """
    Vista para el registro de nuevos usuarios
    PBI-19: Implementa los criterios de aceptación para registro
    """
    if request.method == 'POST':
        form = RegistroForm(request.POST)
        if form.is_valid():
            try:
                # Crear el usuario usando el formulario
                user = form.save()
                
                # Mensaje de éxito
                messages.success(
                    request, 
                    f'¡Bienvenido {user.username}! Tu cuenta ha sido creada exitosamente. '
                    'Ahora puedes iniciar sesión.'
                )
                
                # Redirigir al login después del registro exitoso
                return redirect('login')
                
            except Exception as e:
                messages.error(
                    request, 
                    f'Error al crear la cuenta: {str(e)}. Por favor intenta nuevamente.'
                )
        else:
            # Si el formulario tiene errores, se mostrarán automáticamente
            messages.error(
                request, 
                'Por favor corrige los errores en el formulario.'
            )
    else:
        # GET request - mostrar formulario vacío
        form = RegistroForm()
    
    return render(request, 'registration/registro.html', {
        'form': form,
        'title': 'Registro de Usuario'
    })


#######################################################################

PRECIOS_FORMATO = {
    '2D': 4.00,
    '3D': 6.00,
    'IMAX': 8.00
}

@csrf_exempt
def asientos(request, pelicula_id=None):

    # ========== VALIDACIÓN INICIAL ==========
    pelicula = get_object_or_404(Pelicula, pk=pelicula_id) if pelicula_id else None
    if not pelicula:
        messages.error(request, "No se ha seleccionado ninguna película")
        return redirect('index')

    # ========== MANEJO DE FECHA ==========
    fecha_str = request.GET.get('fecha', '') or request.POST.get('fecha', '')
    ahora_naive = datetime.now()
    
    try:
        fecha_seleccionada = datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else ahora_naive.date()
    except ValueError:
        fecha_seleccionada = ahora_naive.date()

    # ========== FUNCIONES DISPONIBLES ==========
    funciones = Funcion.objects.filter(
        pelicula=pelicula,
        activa=True,
        fecha_inicio__lte=fecha_seleccionada
    ).order_by('horario')

    # Filtrar funciones vigentes considerando duración y hora actual
    funciones_vigentes = []
    for funcion in funciones:
        fecha_fin_funcion = funcion.fecha_inicio + timedelta(weeks=funcion.semanas) - timedelta(days=1)
        
        if funcion.fecha_inicio <= fecha_seleccionada <= fecha_fin_funcion:
            if fecha_seleccionada == ahora_naive.date():
                try:
                    hora_funcion = datetime.strptime(funcion.horario, '%H:%M').time()
                    datetime_funcion = datetime.combine(fecha_seleccionada, hora_funcion)
                    if datetime_funcion > (ahora_naive - timedelta(minutes=10)):
                        funciones_vigentes.append(funcion)
                except ValueError:
                    funciones_vigentes.append(funcion)
            else:
                funciones_vigentes.append(funcion)

    if not funciones_vigentes:
        messages.error(
            request, 
            f"No hay funciones disponibles para esta película el {fecha_seleccionada.strftime('%d/%m/%Y')}"
        )
        return redirect('index')

    # ========== FUNCIÓN ACTUAL ==========
    funcion_actual_id = request.POST.get('funcion_id') or request.GET.get('funcion_id')
    funcion_actual = None
    
    if funcion_actual_id:
        funcion_actual = next(
            (f for f in funciones_vigentes if str(f.id) == str(funcion_actual_id)), 
            None
        )
    
    if not funcion_actual and funciones_vigentes:
        funcion_actual = funciones_vigentes[0]

    # ========== ASIENTOS OCUPADOS (POR SALA + HORARIO + FECHA) ==========
    asientos_ocupados = []
    if funcion_actual:
        reservas_existentes = Reserva.objects.filter(
            pelicula=pelicula,
            sala=str(funcion_actual.sala),
            horario=funcion_actual.horario,
            fecha_funcion=fecha_seleccionada,
            estado__in=['RESERVADO', 'CONFIRMADO']
        )
        
        for reserva in reservas_existentes:
            asientos_ocupados.extend(reserva.get_asientos_list())
        
        asientos_ocupados = list(set(asientos_ocupados))
        print(f"🎬 Asientos ocupados para Sala {funcion_actual.sala}, Horario {funcion_actual.horario}, Fecha {fecha_seleccionada}: {asientos_ocupados}")

    # ========== DATOS DE SELECCIÓN ==========
    asientos_seleccionados = request.POST.getlist("asientos_list")
    cantidad_boletos = len(asientos_seleccionados)

    # Formato y precio de la función actual
    formato_actual = funcion_actual.get_formato_sala() if funcion_actual else '2D'
    precio_funcion_actual = PRECIOS_FORMATO.get(formato_actual, 4.00)

    # ========== APLICAR CUPÓN ==========
    codigo_cupon = request.POST.get("codigo_cupon", "").strip()
    descuento_porcentaje = 0
    mensaje_cupon = ""
    
    if codigo_cupon:
        cupon = CodigoDescuento.objects.filter(
            codigo__iexact=codigo_cupon, 
            estado=True
        ).first()
        
        if cupon:
            descuento_porcentaje = cupon.porcentaje
            mensaje_cupon = f"✅ Cupón aplicado: {descuento_porcentaje}% de descuento"
        else:
            mensaje_cupon = "❌ Código inválido o inactivo"

    # ========== CÁLCULOS DE PRECIO ==========
    subtotal = cantidad_boletos * precio_funcion_actual
    descuento_monto = subtotal * (float(descuento_porcentaje) / 100)
    total = subtotal - descuento_monto

    # ========== RESPUESTA AJAX ==========
    if request.method == "POST" and request.headers.get("X-Requested-With") == "XMLHttpRequest":
        print(f"📡 AJAX Request - Función: {funcion_actual.id if funcion_actual else 'None'}")
        print(f"📡 Asientos ocupados: {asientos_ocupados}")
        
        return JsonResponse({
            "asientos": asientos_seleccionados,
            "cantidad_boletos": cantidad_boletos,
            "formato": formato_actual,
            "precio_boleto": precio_funcion_actual,
            "subtotal": float(subtotal),
            "descuento": descuento_porcentaje,
            "descuento_monto": float(descuento_monto),
            "total": float(total),
            "asientos_ocupados": asientos_ocupados,
            "funcion_id": funcion_actual.id if funcion_actual else None,
            "sala": str(funcion_actual.sala) if funcion_actual else '',
            "horario": funcion_actual.horario if funcion_actual else ''
        })

    # ========== CONFIRMAR RESERVA Y PAGO ==========
    if request.method == "POST" and request.POST.get("accion") == "reservar":
        print("💳 Iniciando proceso de reserva y pago...")
        print(f"📋 Asientos seleccionados recibidos: {asientos_seleccionados}")
        
        # Datos del cliente
        nombre_cliente = request.POST.get("nombre_cliente", "").strip()
        apellido_cliente = request.POST.get("apellido_cliente", "").strip()
        email = request.POST.get("email", "").strip()
        funcion_id = request.POST.get("funcion_id")

        # Detectar tipo de pago
        usar_metodo_guardado = request.POST.get("usar_metodo_guardado", "false")
        
        # Variables de pago
        numero_tarjeta = ""
        nombre_titular = ""
        fecha_expiracion = ""
        cvv = ""
        guardar_tarjeta = False
        alias_tarjeta = ""

        # ========== VALIDACIONES ==========
        errores = []
        
        if not nombre_cliente:
            errores.append("El nombre es obligatorio")
        if not apellido_cliente:
            errores.append("El apellido es obligatorio")
        if not email or "@" not in email:
            errores.append("Ingrese un email válido")
        if not funcion_id:
            errores.append("Seleccione una función")
        if cantidad_boletos == 0:
            errores.append("Seleccione al menos un asiento")
        
        # ⚠️ CRÍTICO: Validar que asientos_seleccionados no esté vacío
        if not asientos_seleccionados:
            errores.append("No se recibieron asientos seleccionados. Por favor, seleccione sus asientos nuevamente.")
            print("❌ ERROR: asientos_seleccionados está vacío!")
        
        # Validar que los asientos seleccionados no estén ocupados
        asientos_ya_ocupados = [a for a in asientos_seleccionados if a in asientos_ocupados]
        if asientos_ya_ocupados:
            errores.append(f"Los siguientes asientos ya están ocupados: {', '.join(asientos_ya_ocupados)}")
        
        metodo_guardado_usado = None
        
        # ========== VALIDACIÓN DE MÉTODO DE PAGO ==========
        if usar_metodo_guardado != "false":
            try:
                metodo_id = int(usar_metodo_guardado)
                metodo = MetodoPago.objects.get(
                    id=metodo_id, 
                    usuario=request.user, 
                    activo=True
                )
                metodo_guardado_usado = metodo
                
                cvv_guardado = request.POST.get("cvv_guardado", "").strip()
                if not cvv_guardado:
                    errores.append("El CVV es obligatorio para usar un método guardado")
                elif len(cvv_guardado) < 3:
                    errores.append("El CVV debe tener al menos 3 dígitos")
                else:
                    datos_tarjeta = decrypt_card_data(metodo.datos_encriptados)
                    
                    numero_tarjeta = datos_tarjeta.get('numero_tarjeta', '').replace(' ', '').replace('-', '')
                    nombre_titular = datos_tarjeta.get('nombre_titular', '')
                    
                    anio_corto = str(metodo.anio_expiracion)[-2:] if metodo.anio_expiracion else '00'
                    mes_formateado = str(metodo.mes_expiracion).zfill(2) if metodo.mes_expiracion else '00'
                    fecha_expiracion = f"{mes_formateado}/{anio_corto}"
                    cvv = cvv_guardado
                    
                    if metodo.esta_expirada():
                        errores.append("El método de pago seleccionado está expirado. Por favor, actualízalo.")
                        
            except (ValueError, MetodoPago.DoesNotExist):
                errores.append("Método de pago no válido")
        
        else:
            numero_tarjeta = request.POST.get("numero_tarjeta", "").strip().replace(" ", "")
            nombre_titular = request.POST.get("nombre_titular", "").strip()
            fecha_expiracion = request.POST.get("fecha_expiracion", "").strip()
            cvv = request.POST.get("cvv", "").strip()
            guardar_tarjeta = request.POST.get("guardar_tarjeta") == "on"
            alias_tarjeta = request.POST.get("alias_tarjeta", "").strip()
            
            if not numero_tarjeta:
                errores.append("El número de tarjeta es obligatorio")
            elif len(numero_tarjeta) < 13:
                errores.append("El número de tarjeta es inválido")
                
            if not nombre_titular:
                errores.append("El nombre del titular es obligatorio")
                
            if not fecha_expiracion:
                errores.append("La fecha de expiración es obligatoria")
            elif '/' not in fecha_expiracion:
                errores.append("Formato de fecha de expiración inválido (use MM/YY)")
                
            if not cvv:
                errores.append("El CVV es obligatorio")
            elif len(cvv) < 3:
                errores.append("El CVV debe tener al menos 3 dígitos")
            
            if guardar_tarjeta and not alias_tarjeta:
                errores.append("Debes proporcionar un nombre para guardar la tarjeta")

        # ========== SI HAY ERRORES, MOSTRARLOS ==========
        if errores:
            for error in errores:
                messages.error(request, error)
            
            funciones_con_precios = []
            for funcion in funciones_vigentes:
                formato = funcion.get_formato_sala()
                precio = PRECIOS_FORMATO.get(formato, 4.00)
                funciones_con_precios.append({
                    "funcion": funcion,
                    "formato": formato,
                    "precio": precio
                })

            metodos_guardados = []
            if request.user.is_authenticated:
                metodos_guardados = MetodoPago.objects.filter(
                    usuario=request.user,
                    activo=True
                ).order_by('-es_predeterminado', '-fecha_creacion')

            context = {
                "pelicula": pelicula,
                "asientos_ocupados": asientos_ocupados,
                "funciones_vigentes": funciones_vigentes,
                "funciones_con_precios": funciones_con_precios,
                "funcion_actual": funcion_actual,
                "fecha_seleccionada": fecha_seleccionada,
                "codigo_reserva": request.session.get("codigo_reserva"),
                "precio_funcion_actual": precio_funcion_actual,
                "formato_actual": formato_actual,
                "asientos_seleccionados": asientos_seleccionados,
                "cantidad_boletos": cantidad_boletos,
                "subtotal": subtotal,
                "descuento_porcentaje": descuento_porcentaje,
                "descuento_monto": descuento_monto,
                "total": total,
                "mensaje_cupon": mensaje_cupon,
                "nombre_cliente": nombre_cliente,
                "apellido_cliente": apellido_cliente,
                "email": email,
                "codigo_cupon": codigo_cupon,
                "metodos_guardados": metodos_guardados,
            }
            return render(request, "asientos.html", context)

        # ========== PROCESAR PAGO Y RESERVA ==========
        try:
            print("🔄 Iniciando transacción atómica...")
            print(f"📍 Asientos a guardar: {asientos_seleccionados}")
            
            funcion = get_object_or_404(Funcion, id=funcion_id)
            formato_funcion = funcion.get_formato_sala()
            print(f"🎬 Formato obtenido de la función: [{formato_funcion}] (tipo: {type(formato_funcion)})")
            
            # Validar que formato no esté vacío
            if not formato_funcion or formato_funcion.strip() == "":
                formato_funcion = "2D"  # Default
                print(f"⚠️ Formato vacío, usando default: {formato_funcion}")
            
            precio_por_boleto = PRECIOS_FORMATO.get(formato_funcion, 4.00)

            # Recalcular totales
            subtotal = cantidad_boletos * precio_por_boleto
            descuento_monto = subtotal * (float(descuento_porcentaje) / 100)
            precio_total = subtotal - descuento_monto

            # ========== VALIDAR DISPONIBILIDAD FINAL ==========
            with transaction.atomic():
                reservas_conflicto = Reserva.objects.filter(
                    pelicula=pelicula,
                    sala=str(funcion.sala),
                    horario=funcion.horario,
                    fecha_funcion=fecha_seleccionada,
                    estado__in=['RESERVADO', 'CONFIRMADO']
                ).select_for_update()  # ⚠️ Bloquear para evitar race conditions
                
                asientos_conflicto = []
                for res in reservas_conflicto:
                    asientos_conflicto.extend(res.get_asientos_list())
                
                asientos_duplicados = [a for a in asientos_seleccionados if a in asientos_conflicto]
                if asientos_duplicados:
                    raise Exception(f"Los asientos {', '.join(asientos_duplicados)} ya fueron reservados por otro usuario")

                # ========== PROCESAR PAGO SIMULADO ==========
                print(f"🔍 DEBUG - numero_tarjeta: [{numero_tarjeta}] (tipo: {type(numero_tarjeta)})")
                print(f"🔍 DEBUG - nombre_titular: [{nombre_titular}]")
                print(f"🔍 DEBUG - fecha_expiracion: [{fecha_expiracion}]")
                print(f"🔍 DEBUG - cvv: [{cvv}]")
                
                datos_tarjeta = {
                    'numero': numero_tarjeta,
                    'mes_expiracion': int(fecha_expiracion.split('/')[0]),
                    'anio_expiracion': int('20' + fecha_expiracion.split('/')[1]),
                    'cvv': cvv,
                    'nombre_titular': nombre_titular
                }
                
                print(f"💳 Procesando pago de ${precio_total:.2f}...")
                resultado_pago = simular_pago(
                    datos_tarjeta=datos_tarjeta,
                    monto=float(precio_total)
                )

                if not resultado_pago["exitoso"]:
                    error_msg = resultado_pago.get('error_message', 'Pago rechazado')
                    print(f"❌ Pago rechazado: {error_msg}")
                    raise Exception(f"Pago rechazado: {error_msg}")

                print("✅ Pago aprobado!")

                # ========== CREAR CÓDIGO DE RESERVA ==========
                codigo_reserva = "".join(random.choices(string.ascii_uppercase + string.digits, k=8))
                
                asientos_finales_unicos = sorted(list(set(asientos_seleccionados)))

                # 2. Crear el string final
                asientos_str = ",".join(asientos_finales_unicos)
                print(f"💾 Guardando reserva con asientos: '{asientos_str}'")
                
                reserva = Reserva(
                    pelicula=pelicula,
                    nombre_cliente=nombre_cliente,
                    apellido_cliente=apellido_cliente,
                    email=email,
                    formato=formato_funcion,
                    sala=str(funcion.sala),
                    horario=funcion.horario,
                    fecha_funcion=fecha_seleccionada,
                    asientos=asientos_str,  # ✅ String separado por comas
                    cantidad_boletos=cantidad_boletos,
                    precio_total=precio_total,
                    estado="CONFIRMADO",
                    usuario=request.user,
                    pago_completado=True,
                    fecha_pago=timezone.now(),
                    codigo_reserva=codigo_reserva
                )
                reserva.save()
                print(f"✅ Reserva guardada - ID: {reserva.id}, Código: {codigo_reserva}")
                print(f"✅ Asientos en BD: '{reserva.asientos}'")

                # ========== CREAR REGISTRO DE PAGO ==========
                pago = Pago(
                    reserva=reserva,  # ✅ Ahora sí existe la reserva
                    monto=precio_total,
                    metodo_pago="TARJETA",
                    estado_pago="APROBADO",
                    numero_transaccion=resultado_pago.get("numero_transaccion") or "",
                    detalles_pago={
                        "numero_tarjeta_enmascarado": resultado_pago.get("numero_tarjeta_enmascarado", ""),
                        "nombre_titular": nombre_titular,
                        "tipo_tarjeta": resultado_pago.get("tipo_tarjeta", ""),
                    },
                    metodo_pago_guardado=metodo_guardado_usado
                )
                pago.save()
                print(f"💾 Pago guardado - ID: {pago.id}")

                # ========== REGISTRAR VENTA ==========
                print(f"📊 Creando Venta con formato: [{reserva.formato}] (tipo: {type(reserva.formato)})")
                from django.utils import timezone
                Venta.objects.create(
                    pelicula=reserva.pelicula,
                    sala=reserva.sala,
                    fecha=fecha_seleccionada,
                    cantidad_boletos=reserva.cantidad_boletos,
                    total_venta=reserva.precio_total,
                    formato=reserva.formato if reserva.formato else "2D",  # Asegurar que no sea None/vacío
                    fecha_venta=timezone.now().date()  # Fecha actual de la venta
                )
                print("📊 Venta registrada")

                # ========== GUARDAR MÉTODO DE PAGO (SI SE SOLICITÓ) ==========
                if guardar_tarjeta and request.user.is_authenticated:
                    try:
                        datos_encriptados = encrypt_card_data_full(
                            numero_tarjeta=numero_tarjeta,
                            nombre_titular=nombre_titular,
                            fecha_expiracion=fecha_expiracion
                        )
                        
                        tipo_tarjeta = resultado_pago.get("tipo_tarjeta", "OTRA")
                        
                        metodo_existente = MetodoPago.objects.filter(
                            usuario=request.user,
                            alias=alias_tarjeta
                        ).first()
                        
                        if metodo_existente:
                            metodo_existente.datos_encriptados = datos_encriptados
                            metodo_existente.ultimos_4_digitos = numero_tarjeta[-4:]
                            metodo_existente.tipo_tarjeta = tipo_tarjeta
                            metodo_existente.mes_expiracion = int(fecha_expiracion.split('/')[0])
                            metodo_existente.anio_expiracion = int('20' + fecha_expiracion.split('/')[1])
                            metodo_existente.nombre_titular = nombre_titular
                            metodo_existente.activo = True
                            metodo_existente.save()
                            messages.success(request, f"✓ Método de pago '{alias_tarjeta}' actualizado")
                        else:
                            MetodoPago.objects.create(
                                usuario=request.user,
                                tipo='TARJETA',
                                alias=alias_tarjeta,
                                datos_encriptados=datos_encriptados,
                                ultimos_4_digitos=numero_tarjeta[-4:],
                                tipo_tarjeta=tipo_tarjeta,
                                mes_expiracion=int(fecha_expiracion.split('/')[0]),
                                anio_expiracion=int('20' + fecha_expiracion.split('/')[1]),
                                nombre_titular=nombre_titular,
                                es_predeterminado=False,
                                activo=True
                            )
                            messages.success(request, f"✓ Método de pago '{alias_tarjeta}' guardado")
                    except Exception as e:
                        print(f"⚠️ Error guardando método de pago: {str(e)}")
                        messages.warning(request, f"La reserva fue exitosa pero no se pudo guardar el método de pago: {str(e)}")

            # ========== GENERAR PDF Y ENVIAR EMAIL (FUERA DE TRANSACCIÓN) ==========
            # ⚠️ IMPORTANTE: Esto va DESPUÉS del commit de la transacción
            try:
                print("📧 Generando PDF y enviando email...")
                pdf_buffer = generar_pdf_reserva(reserva)
                subject = f"Confirmación de Reserva - Código {reserva.codigo_reserva}"
                body = (
                    f"Hola {reserva.nombre_cliente},<br><br>"
                    f"Tu reserva para la película '<strong>{reserva.pelicula.nombre}</strong>' ha sido confirmada.<br><br>"
                    f"<strong>Detalles de tu reserva:</strong><br>"
                    f"🎫 Código de reserva: <strong>{reserva.codigo_reserva}</strong><br>"
                    f"📅 Fecha: {fecha_seleccionada.strftime('%d/%m/%Y')}<br>"
                    f"🕐 Horario: {funcion.get_horario_display()}<br>"
                    f"🎬 Sala: {funcion.sala}<br>"
                    f"📽️ Formato: {formato_funcion}<br>"
                    f"💺 Asientos: {', '.join(asientos_seleccionados)}<br><br>"
                    f"<strong>Resumen de pago:</strong><br>"
                    f"Subtotal: ${subtotal:.2f}<br>"
                    f"Descuento: -${descuento_monto:.2f}<br>"
                    f"<strong>Total: ${precio_total:.2f}</strong><br>"
                    f"💳 Transacción: {pago.numero_transaccion}<br><br>"
                    "Adjunto encontrarás tu ticket en formato PDF.<br><br>"
                    "¡Gracias por elegir CineDot! 🎬🍿"
                )
                
                send_brevo_email(
                    to_emails=[reserva.email],
                    subject=subject,
                    html_content=body,
                    attachments=[(
                        f"ticket_{reserva.codigo_reserva}.pdf", 
                        pdf_buffer.getvalue(), 
                        "application/pdf"
                    )]
                )
                print("✅ Email enviado con éxito")
            except Exception as e:
                print(f"⚠️ Error al enviar email (no afecta la reserva): {str(e)}")
                import traceback
                traceback.print_exc()
                # No fallar la reserva si falla el email

            # ========== ÉXITO ==========
            request.session["codigo_reserva"] = reserva.codigo_reserva
            messages.success(
                request, 
                f"🎉 ¡Pago y reserva exitosos! Código: {reserva.codigo_reserva}. Revisa tu email para obtener tu ticket."
            )
            
            print(f"✅ Proceso completado exitosamente - Código: {codigo_reserva}")
            
            return redirect(
                f"{reverse('asientos', args=[pelicula.id])}?fecha={fecha_seleccionada.strftime('%Y-%m-%d')}"
            )
                
        except Exception as e:
            print(f"❌ ERROR GENERAL en transacción: {str(e)}")
            import traceback
            traceback.print_exc()  # ⚠️ Ver traceback completo
            
            if "Pago rechazado" in str(e):
                messages.error(request, str(e))
            else:
                messages.error(request, f"Error al procesar la reserva: {str(e)}")

    # ========== RENDER NORMAL (GET O POST SIN RESERVAR) ==========
    funciones_con_precios = []
    for funcion in funciones_vigentes:
        formato = funcion.get_formato_sala()
        precio = PRECIOS_FORMATO.get(formato, 4.00)
        funciones_con_precios.append({
            "funcion": funcion,
            "formato": formato,
            "precio": precio
        })

    metodos_guardados = []
    if request.user.is_authenticated:
        metodos_guardados = MetodoPago.objects.filter(
            usuario=request.user,
            activo=True
        ).order_by('-es_predeterminado', '-fecha_creacion')

    context = {
        "pelicula": pelicula,
        "asientos_ocupados": asientos_ocupados,
        "funciones_vigentes": funciones_vigentes,
        "funciones_con_precios": funciones_con_precios,
        "funcion_actual": funcion_actual,
        "fecha_seleccionada": fecha_seleccionada,
        "codigo_reserva": request.session.get("codigo_reserva"),
        "precio_funcion_actual": precio_funcion_actual,
        "formato_actual": formato_actual,
        "asientos_seleccionados": asientos_seleccionados,
        "cantidad_boletos": cantidad_boletos,
        "subtotal": subtotal,
        "descuento_porcentaje": descuento_porcentaje,
        "descuento_monto": descuento_monto,
        "total": total,
        "mensaje_cupon": mensaje_cupon,
        "nombre_cliente": request.POST.get("nombre_cliente", ""),
        "apellido_cliente": request.POST.get("apellido_cliente", ""),
        "email": request.POST.get("email", ""),
        "codigo_cupon": codigo_cupon,
        "metodos_guardados": metodos_guardados,
    }
    
    return render(request, "asientos.html", context)
#################################################################
#################################################################
@csrf_exempt
@require_POST
def aplicar_descuento_ajax(request):
    """
    Función que valida un código de descuento y guarda el porcentaje en la sesión.
    Responde con JSON.
    """
    if request.method == 'POST':
        codigo = request.POST.get('codigo', '').strip().upper()
        
        if not codigo:
            return JsonResponse({'success': False, 'mensaje': 'Ingrese un código de descuento.'})

        try:
            cupon = CodigoDescuento.objects.get(codigo=codigo, estado=True)

            # Aplicacion de el descuento y guardarlo en la sesión
            descuento = cupon.porcentaje
            
            descuento = float(cupon.porcentaje or 0)
            request.session['descuento_porcentaje'] = descuento
            
            mensaje = f'Cupón "{codigo}" aplicado: ¡{descuento}% de descuento!'
            request.session['mensaje_cupon'] = mensaje
            
            return JsonResponse({'success': True, 'descuento_porcentaje': descuento, 'mensaje': mensaje})

        except CodigoDescuento.DoesNotExist:
            return JsonResponse({'success': False, 'mensaje': 'Código no válido o inactivo.'})

        except Exception as e:
            return JsonResponse({'success': False, 'mensaje': f'Error interno: {str(e)}'})

    return JsonResponse({'success': False, 'mensaje': 'Método no permitido.'})
#################################################################
#################################################################
def registrar_cupon(request):
  
    if request.method == 'POST':
        
        form = CodigoDescuentoForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect('registrar_cupon') 
    
    else:
        form = CodigoDescuentoForm()

    cupones_registrados = CodigoDescuento.objects.all().order_by('-id')

    contexto = {
        'form': form,
        'cupones': cupones_registrados 
    }

    return render(request, 'registrar_cupon.html', contexto)
    
def eliminar_cupon(request, pk):
   
    cupon = get_object_or_404(CodigoDescuento, pk=pk)
    cupon.delete() 
    return redirect('registrar_cupon')

def modificar_cupon(request, pk):
    cupon = get_object_or_404(CodigoDescuento, pk=pk)
    
    if request.method == 'POST':
        form = CodigoDescuentoForm(request.POST, instance=cupon)
        if form.is_valid():
            form.save()
            return redirect('registrar_cupon') 
    else:
        form = CodigoDescuentoForm(instance=cupon)
        
    cupones_registrados = CodigoDescuento.objects.all().order_by('-id')
    
    contexto = {
        'form': form,
        'cupones': cupones_registrados,
        'es_edicion': True, # Sirve para identificar si se esta modificandi
        'cupon_id': pk
    }
    
    return render(request, 'registrar_cupon.html', contexto)

##################################################################################
##################################################################################

@csrf_exempt
def administrar_salas(request):
    """
    Vista para administrar asientos de salas por película, fecha, horario y sala.
    Solo muestra películas con funciones activas en cartelera.
    """
    
    # ========== OBTENER FECHA SELECCIONADA ==========
    ahora_naive = datetime.now()
    hoy = ahora_naive.date()
    
    fecha_str = request.GET.get('fecha', '') or request.POST.get('fecha', '')
    try:
        fecha_seleccionada = datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else hoy
    except ValueError:
        fecha_seleccionada = hoy
    
    # ========== GENERAR FECHAS DISPONIBLES (5 DÍAS) ==========
    fechas_disponibles = []
    nombres_dias_es = {
        'Monday': 'Lun', 'Tuesday': 'Mar', 'Wednesday': 'Mié',
        'Thursday': 'Jue', 'Friday': 'Vie', 'Saturday': 'Sáb', 'Sunday': 'Dom'
    }
    
    for i in range(5):
        dia = hoy + timedelta(days=i)
        fechas_disponibles.append({
            'fecha': dia,
            'nombre': nombres_dias_es[dia.strftime('%A')],
            'formato': dia.strftime('%d/%m'),
            'es_hoy': dia == hoy
        })
    
    # ========== OBTENER PELÍCULAS EN CARTELERA ==========
    # Solo películas con funciones activas que estén vigentes en la fecha seleccionada
    funciones_activas = Funcion.objects.filter(
        activa=True,
        fecha_inicio__lte=fecha_seleccionada
    ).select_related('pelicula')
    
    # Filtrar funciones que no hayan expirado
    peliculas_ids = set()
    for funcion in funciones_activas:
        fecha_fin = funcion.fecha_inicio + timedelta(weeks=funcion.semanas) - timedelta(days=1)
        if fecha_fin >= fecha_seleccionada:
            peliculas_ids.add(funcion.pelicula.id)
    
    peliculas = Pelicula.objects.filter(id__in=peliculas_ids).order_by('nombre')
    
    # ========== PROCESAR PELÍCULA SELECCIONADA ==========
    pelicula_id = request.GET.get('pelicula') or request.POST.get('pelicula')
    pelicula = None
    funciones_vigentes = []
    combinaciones = []
    combo_actual = request.GET.get('combo') or request.POST.get('combo')
    asientos_ocupados = []
    
    if pelicula_id:
        pelicula = Pelicula.objects.filter(id=pelicula_id).first()
        
        if pelicula:
            # Obtener funciones vigentes para esta película en la fecha seleccionada
            funciones = Funcion.objects.filter(
                pelicula=pelicula,
                activa=True,
                fecha_inicio__lte=fecha_seleccionada
            ).order_by('horario', 'sala')
            
            for funcion in funciones:
                fecha_fin = funcion.fecha_inicio + timedelta(weeks=funcion.semanas) - timedelta(days=1)
                
                if fecha_fin >= fecha_seleccionada:
                    # Si es hoy, verificar que no haya pasado el horario
                    if fecha_seleccionada == hoy:
                        try:
                            hora_funcion = datetime.strptime(funcion.horario, '%H:%M').time()
                            datetime_funcion = datetime.combine(fecha_seleccionada, hora_funcion)
                            if datetime_funcion > (ahora_naive - timedelta(minutes=10)):
                                funciones_vigentes.append(funcion)
                        except ValueError:
                            funciones_vigentes.append(funcion)
                    else:
                        funciones_vigentes.append(funcion)
            
            # Crear combinaciones: "14:30 - Sala 1 (2D)"
            for funcion in funciones_vigentes:
                formato = funcion.get_formato_sala()
                combo_str = f"{funcion.horario} - Sala {funcion.sala} ({formato})"
                combinaciones.append(combo_str)
            
            # ========== OBTENER ASIENTOS OCUPADOS ==========
            if combo_actual:
                try:
                    # Parsear: "14:30 - Sala 1 (2D)"
                    partes = combo_actual.split(" - Sala ")
                    horario_sel = partes[0].strip()
                    
                    sala_y_formato = partes[1]
                    sala_sel = sala_y_formato.split(" (")[0].strip()
                    
                    # Buscar reservas para esta combinación específica
                    reservas = Reserva.objects.filter(
                        pelicula=pelicula,
                        sala=sala_sel,
                        horario=horario_sel,
                        fecha_funcion=fecha_seleccionada,
                        estado__in=['RESERVADO', 'CONFIRMADO']
                    )
                    
                    for reserva in reservas:
                        asientos_ocupados.extend(reserva.get_asientos_list())
                    
                    asientos_ocupados = list(set(asientos_ocupados))
                    
                    print(f"🎬 Admin Salas - Película: {pelicula.nombre}")
                    print(f"📅 Fecha: {fecha_seleccionada}")
                    print(f"🕐 Horario: {horario_sel}")
                    print(f"🎭 Sala: {sala_sel}")
                    print(f"💺 Asientos ocupados: {asientos_ocupados}")
                    
                except (ValueError, IndexError) as e:
                    print(f"❌ Error parseando combo: {e}")
    
    # ========== MANEJO DE ACCIONES POST ==========
    if request.method == 'POST':
        if not combo_actual or not pelicula:
            return JsonResponse({'success': False, 'error': 'Selecciona película y horario'})
        
        try:
            # Parsear combo actual
            partes = combo_actual.split(" - Sala ")
            horario_sel = partes[0].strip()
            sala_sel = partes[1].split(" (")[0].strip()
            
            # ========== RESTABLECER TODOS LOS ASIENTOS ==========
            if request.POST.get('restablecer') == 'true':
                reservas_eliminadas = Reserva.objects.filter(
                    pelicula=pelicula,
                    sala=sala_sel,
                    horario=horario_sel,
                    fecha_funcion=fecha_seleccionada,
                    estado__in=['RESERVADO', 'CONFIRMADO']
                )
                
                count = reservas_eliminadas.count()
                reservas_eliminadas.delete()
                
                print(f"✅ Restablecidos todos los asientos: {count} reservas eliminadas")
                return JsonResponse({
                    'success': True,
                    'mensaje': f'{count} reservas eliminadas'
                })
            
            # ========== ELIMINAR ASIENTO INDIVIDUAL ==========
            elif request.POST.get('eliminar_asiento'):
                asiento = request.POST.get('eliminar_asiento').strip()
                
                reservas = Reserva.objects.filter(
                    pelicula=pelicula,
                    sala=sala_sel,
                    horario=horario_sel,
                    fecha_funcion=fecha_seleccionada,
                    estado__in=['RESERVADO', 'CONFIRMADO']
                )
                
                eliminado = False
                for reserva in reservas:
                    asientos = reserva.get_asientos_list()
                    
                    if asiento in asientos:
                        asientos.remove(asiento)
                        
                        if asientos:
                            # Actualizar cantidad de boletos
                            reserva.asientos = ",".join(asientos)
                            reserva.cantidad_boletos = len(asientos)
                            
                            # Recalcular precio
                            from .utils import PRECIOS_FORMATO
                            formato_reserva = reserva.formato
                            precio_boleto = PRECIOS_FORMATO.get(formato_reserva, 4.00)
                            reserva.precio_total = precio_boleto * reserva.cantidad_boletos
                            
                            reserva.save()
                            print(f"✅ Asiento {asiento} eliminado. Quedan: {asientos}")
                        else:
                            # Si no quedan asientos, eliminar reserva completa
                            reserva.delete()
                            print(f"✅ Reserva eliminada (era el último asiento)")
                        
                        eliminado = True
                        break
                
                if eliminado:
                    return JsonResponse({
                        'success': True,
                        'mensaje': f'Asiento {asiento} eliminado'
                    })
                else:
                    return JsonResponse({
                        'success': False,
                        'error': f'Asiento {asiento} no encontrado'
                    })
        
        except Exception as e:
            print(f"❌ Error en POST: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    # ========== CONTEXTO PARA TEMPLATE ==========
    context = {
        'peliculas': peliculas,
        'pelicula': pelicula,
        'funciones_vigentes': funciones_vigentes,
        'combinaciones': combinaciones,
        'combo_actual': combo_actual,
        'asientos_ocupados': asientos_ocupados,
        'fecha_seleccionada': fecha_seleccionada,
        'fechas_disponibles': fechas_disponibles,
    }
    
    return render(request, 'administrar_salas.html', context)


################################################################


def generar_pdf_reserva(reserva):
    buffer = BytesIO()
    
    # Obtener fecha y hora actual del sistema
    ahora = datetime.now()
    fecha_emision = ahora.strftime('%d/%m/%Y %H:%M:%S')

    # Configurar el documento
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                          rightMargin=72, leftMargin=72,
                          topMargin=72, bottomMargin=72)
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Crear estilos personalizados
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Title'],
        fontSize=20,
        leading=24,
        spaceAfter=20,
        alignment=1,
        textColor=colors.HexColor('#2c3e50')
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Heading2'],
        fontSize=14,
        leading=18,
        spaceAfter=12,
        alignment=1,
        textColor=colors.HexColor('#3498db')
    )
    
    info_style = ParagraphStyle(
        'Info',
        parent=styles['Normal'],
        fontSize=12,
        leading=15,
        spaceAfter=8
    )
    
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=10,
        leading=12,
        textColor=colors.grey,
        alignment=1
    )
    
    # Contenido del PDF
    elements = []
    
    # Logo
    logo_path = os.path.abspath(os.path.join(settings.BASE_DIR, 'myapp', 'static', 'imagenes', 'cine.png'))
    if os.path.exists(logo_path):
        pil_logo = PILImage.open(logo_path)
        original_width, original_height = pil_logo.size

        # Convertir a pulgadas
        width_inch = original_width / 96
        height_inch = original_height / 96

        # Escalar a la mitad
        scale = 0.5
        final_width = width_inch * inch * scale
        final_height = height_inch * inch * scale

        # Insertar logo sin margen superior
        logo = Image(logo_path, width=final_width, height=final_height)
        elements.append(logo)

        # Espacio opcional debajo del logo
        elements.append(Spacer(1, 0.1 * inch))



    # Título
    elements.append(Paragraph("CineDot", title_style))
    elements.append(Paragraph("Ticket de Reserva", subtitle_style))
    elements.append(Spacer(1, 20))
    
    # Información de emisión
    emision_data = [
        [Paragraph("<b>Fecha de emisión:</b>", info_style), 
         Paragraph(fecha_emision, info_style)]
    ]
    
    emision_table = Table(emision_data, colWidths=[1.5*inch, 4*inch])
    emision_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (0,0), (0,-1), 'RIGHT'),
        ('ALIGN', (1,0), (1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 12),
        ('BOTTOMPADDING', (0,0), (-1,-1), 12),
    ]))
    elements.append(emision_table)
    
    # Información del cliente
    cliente_data = [
        [Paragraph("<b>Cliente:</b>", info_style), 
         Paragraph(f"{reserva.nombre_cliente} {reserva.apellido_cliente}", info_style)],
        [Paragraph("<b>Email:</b>", info_style), 
         Paragraph(reserva.email, info_style)],
    ]
    
    cliente_table = Table(cliente_data, colWidths=[1.5*inch, 4*inch])
    cliente_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (0,0), (0,-1), 'RIGHT'),
        ('ALIGN', (1,0), (1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 12),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(cliente_table)
    elements.append(Spacer(1, 20))
    
    # Información de la reserva
    reserva_data = [
        [Paragraph("<b>Código de reserva:</b>", info_style), 
         Paragraph(reserva.codigo_reserva, info_style)],
        [Paragraph("<b>Fecha de reserva:</b>", info_style), 
         Paragraph(reserva.fecha_reserva.strftime('%d/%m/%Y %H:%M'), info_style)],
        [Paragraph("<b>Película:</b>", info_style), 
         Paragraph(reserva.pelicula.nombre, info_style)],
        [Paragraph("<b>Formato:</b>", info_style), 
         Paragraph(reserva.get_formato_display(), info_style)],
        [Paragraph("<b>Sala:</b>", info_style), 
         Paragraph(reserva.sala, info_style)],
        [Paragraph("<b>Horario:</b>", info_style), 
         Paragraph(reserva.horario, info_style)],
        [Paragraph("<b>Asientos:</b>", info_style), 
         Paragraph(reserva.asientos, info_style)],
        [Paragraph("<b>Cantidad de boletos:</b>", info_style), 
         Paragraph(str(reserva.cantidad_boletos), info_style)],
        [Paragraph("<b>Total:</b>", info_style), 
         Paragraph(f"${reserva.precio_total:.2f}", info_style)],
    ]
    
    reserva_table = Table(reserva_data, colWidths=[1.5*inch, 4*inch])
    reserva_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (0,0), (0,-1), 'RIGHT'),
        ('ALIGN', (1,0), (1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 12),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-2), 1, colors.lightgrey),
        ('GRID', (0,-1), (-1,-1), 1, colors.HexColor('#3498db')),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#f8f9fa')),
    ]))
    elements.append(reserva_table)
    elements.append(Spacer(1, 30))
        # Código QR
    qr_image = generar_qr(reserva)
    elements.append(Paragraph("Escanee este código para validar su ticket", info_style))
    elements.append(qr_image)
    elements.append(Spacer(1, 20))
    
    # Mensaje de agradecimiento
    elements.append(Paragraph("Presente este ticket en la entrada del cine", footer_style))
    elements.append(Paragraph("¡Gracias por su preferencia!", footer_style))


    
    # Construir el PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer






def descargar_ticket(request, codigo_reserva):
    reserva = get_object_or_404(Reserva, codigo_reserva=codigo_reserva)
    pdf_buffer = generar_pdf_reserva(reserva)

    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ticket_{reserva.codigo_reserva}.pdf"'
    return response
##########################################################################


def generar_qr(reserva):
    url = f"https://system-design.onrender.com/validaQR/{reserva.codigo_reserva}/"
    qr = qrcode.make(url)
    buffer = BytesIO()
    qr.save(buffer, format='PNG')
    buffer.seek(0)
    return RLImage(buffer, width=1.5*inch, height=1.5*inch)
########################################################################################

def validaQR(request, codigo_reserva):
    reserva = get_object_or_404(Reserva, codigo_reserva=codigo_reserva)

    if reserva.usado:
        mensaje = "❌ Ticket inválido: ya fue utilizado."
        valido = False
    else:
        reserva.usado = True
        reserva.save()
        mensaje = "✅ Ticket válido: bienvenido al cine."
        valido = True

    return render(request, "validaQR.html", {
        "mensaje": mensaje,
        "valido": valido,
        "reserva": reserva,
        "pelicula": reserva.pelicula
    })

################################################################################
# VALORACIONES Y RESEÑAS - PBI-24
################################################################################

def pelicula_detalle(request, pelicula_id):
    """Vista para mostrar detalle de película con valoraciones"""
    pelicula = get_object_or_404(Pelicula, id=pelicula_id)
    
    # Obtener todas las valoraciones de la película
    valoraciones_list = pelicula.valoraciones.all().order_by('-fecha_creacion')
    
    # Paginación de valoraciones
    paginator = Paginator(valoraciones_list, 5)  # 5 valoraciones por página
    page_number = request.GET.get('page')
    valoraciones = paginator.get_page(page_number)
    
    # Verificar si el usuario ya ha valorado esta película
    valoracion_usuario = None
    if request.user.is_authenticated:
        try:
            valoracion_usuario = Valoracion.objects.get(pelicula=pelicula, usuario=request.user)
        except Valoracion.DoesNotExist:
            valoracion_usuario = None
    
    # Convertir géneros para mostrar nombres completos
    pelicula.get_generos_list = convertir_generos(pelicula.generos)
    
    context = {
        'pelicula': pelicula,
        'valoraciones': valoraciones,
        'valoracion_usuario': valoracion_usuario,
        'puede_valorar': request.user.is_authenticated and not valoracion_usuario,
    }
    
    return render(request, 'pelicula_detalle.html', context)


@login_required
def crear_valoracion(request, pelicula_id):
    """Vista para crear o actualizar valoración de una película"""
    pelicula = get_object_or_404(Pelicula, id=pelicula_id)
    
    # Verificar si ya existe una valoración del usuario
    try:
        valoracion_existente = Valoracion.objects.get(pelicula=pelicula, usuario=request.user)
    except Valoracion.DoesNotExist:
        valoracion_existente = None
    
    if request.method == 'POST':
        form = ValoracionForm(
            request.POST, 
            instance=valoracion_existente,
            pelicula=pelicula, 
            usuario=request.user
        )
        
        if form.is_valid():
            valoracion = form.save()
            if valoracion_existente:
                messages.success(request, '¡Tu valoración ha sido actualizada!')
            else:
                messages.success(request, '¡Gracias por tu valoración!')
            
            return redirect('pelicula_detalle', pelicula_id=pelicula.id)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = ValoracionForm(
            instance=valoracion_existente,
            pelicula=pelicula, 
            usuario=request.user
        )
    
    context = {
        'form': form,
        'pelicula': pelicula,
        'valoracion_existente': valoracion_existente,
    }
    
    return render(request, 'crear_valoracion.html', context)


@login_required  
def eliminar_valoracion(request, pelicula_id):
    """Vista para eliminar valoración propia"""
    pelicula = get_object_or_404(Pelicula, id=pelicula_id)
    
    try:
        valoracion = Valoracion.objects.get(pelicula=pelicula, usuario=request.user)
        valoracion.delete()
        messages.success(request, 'Tu valoración ha sido eliminada.')
    except Valoracion.DoesNotExist:
        messages.error(request, 'No se encontró tu valoración.')
    
    return redirect('pelicula_detalle', pelicula_id=pelicula.id)

###############################################################################

def filtrar_peliculas(request):
    """
    Muestra SOLO películas con funciones activas vigentes para la fecha seleccionada
    """
    from django.db.models import Q
    
    genero = request.GET.get('genero', '').strip()
    clasificacion = request.GET.get('clasificacion', '').strip()
    idioma = request.GET.get('idioma', '').strip()

    # ✅ Usar datetime.now() sin zona horaria
    ahora_naive = datetime.now()
    hoy = ahora_naive.date()
    
    # ✅ Obtener fecha seleccionada (si existe)
    fecha_seleccionada_str = request.GET.get('fecha')
    if fecha_seleccionada_str:
        try:
            fecha_seleccionada = datetime.strptime(fecha_seleccionada_str, '%Y-%m-%d').date()
        except ValueError:
            fecha_seleccionada = hoy
    else:
        fecha_seleccionada = hoy

    # ✅ Nombres de meses y días en español
    nombres_meses_es = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto', 
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    
    nombres_dias_es = {
        'Monday': 'Lunes',
        'Tuesday': 'Martes', 
        'Wednesday': 'Miércoles',
        'Thursday': 'Jueves',
        'Friday': 'Viernes',
        'Saturday': 'Sábado',
        'Sunday': 'Domingo'
    }
    
    # ✅ Generar lista de 5 días (hoy + 4 días siguientes)
    dias_disponibles = []
    for i in range(5):
        dia = hoy + timedelta(days=i)
        dia_info = {
            'fecha': dia,
            'nombre_es': nombres_dias_es[dia.strftime('%A')],
            'formato_corto': dia.strftime('%d/%m')
        }
        dias_disponibles.append(dia_info)

    # ✅ Información de fecha seleccionada en español (FORMATO COMPLETO)
    nombre_dia_seleccionado = nombres_dias_es[fecha_seleccionada.strftime('%A')]
    nombre_mes_seleccionado = nombres_meses_es[fecha_seleccionada.month]
    
    # Formato: "Lunes 03 de Noviembre"
    fecha_formateada = f"{nombre_dia_seleccionado} {fecha_seleccionada.day:02d} de {nombre_mes_seleccionado}"

    # ✅ Filtrar funciones vigentes para la fecha seleccionada
    funciones = (
        Funcion.objects.filter(
            activa=True,
            fecha_inicio__lte=fecha_seleccionada,
        )
        .select_related('pelicula')
        .order_by('-pelicula__id', 'horario')
    )

    funciones_filtradas = []
    for funcion in funciones:
        fecha_fin_funcion = funcion.fecha_inicio + timedelta(weeks=funcion.semanas) - timedelta(days=1)
        
        # Verificar que la función esté en el rango de fechas
        if funcion.fecha_inicio <= fecha_seleccionada <= fecha_fin_funcion:
            # Si es hoy, verificar que no haya pasado (margen de 10 minutos)
            if fecha_seleccionada == hoy:
                hora_funcion = datetime.strptime(funcion.horario, '%H:%M').time()
                datetime_funcion = datetime.combine(fecha_seleccionada, hora_funcion)
                
                margen = timedelta(minutes=10)
                if datetime_funcion > (ahora_naive - margen):
                    funciones_filtradas.append(funcion)
            else:
                # Para fechas futuras, incluir todas
                funciones_filtradas.append(funcion)

    # ✅ Agrupar funciones por película
    peliculas_dict = {}
    for funcion in funciones_filtradas:
        pelicula_id = funcion.pelicula.id
        if pelicula_id not in peliculas_dict:
            peliculas_dict[pelicula_id] = {
                'pelicula': funcion.pelicula,
                'funciones': []
            }
        peliculas_dict[pelicula_id]['funciones'].append(funcion)

    # ✅ Convertir a lista de películas
    peliculas_base = [item['pelicula'] for item in peliculas_dict.values()]

    # 🔹 Aplicar filtros del formulario
    peliculas_filtradas = []
    for pelicula in peliculas_base:
        cumple_filtros = True
        
        if genero and genero not in pelicula.generos:
            cumple_filtros = False
        if clasificacion and pelicula.clasificacion != clasificacion:
            cumple_filtros = False
        if idioma and pelicula.idioma != idioma:
            cumple_filtros = False
            
        if cumple_filtros:
            # Agregar funciones del día a la película
            pelicula.funciones_del_dia = peliculas_dict[pelicula.id]['funciones']
            peliculas_filtradas.append(pelicula)

    # 🔹 Enriquecer películas con información adicional
    for pelicula in peliculas_filtradas:
        pelicula.generos_list = pelicula.get_generos_list()
        pelicula.salas_con_formato = pelicula.get_salas_con_formato()

    context = {
        'peliculas': peliculas_filtradas,
        'genero': genero,
        'clasificacion': clasificacion,
        'idioma': idioma,
        'GENERO_CHOICES': Pelicula.GENERO_CHOICES,
        'dias_disponibles': dias_disponibles,
        'fecha_seleccionada': fecha_seleccionada,
        'fecha_formateada': fecha_formateada,
        'hoy': hoy,
    }

    return render(request, 'filtrar.html', context)


###################################################################################

def horarios_por_pelicula(request):
    """
    Muestra SOLO películas con funciones activas vigentes para la fecha seleccionada
    con todos sus horarios y salas disponibles
    """
    # ✅ Usar datetime.now() sin zona horaria
    ahora_naive = datetime.now()
    hoy = ahora_naive.date()

    # ✅ Obtener fecha seleccionada (si existe)
    fecha_seleccionada_str = request.GET.get('fecha')
    if fecha_seleccionada_str:
        try:
            fecha_seleccionada = datetime.strptime(fecha_seleccionada_str, '%Y-%m-%d').date()
        except ValueError:
            fecha_seleccionada = hoy
    else:
        fecha_seleccionada = hoy

    # ✅ Nombres de meses y días en español
    nombres_meses_es = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto', 
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    
    nombres_dias_es = {
        'Monday': 'Lunes',
        'Tuesday': 'Martes', 
        'Wednesday': 'Miércoles',
        'Thursday': 'Jueves',
        'Friday': 'Viernes',
        'Saturday': 'Sábado',
        'Sunday': 'Domingo'
    }
    
    # ✅ Generar lista de 5 días (hoy + 4 días siguientes)
    dias_disponibles = []
    for i in range(5):
        dia = hoy + timedelta(days=i)
        dia_info = {
            'fecha': dia,
            'nombre_es': nombres_dias_es[dia.strftime('%A')],
            'formato_corto': dia.strftime('%d/%m')
        }
        dias_disponibles.append(dia_info)

    # ✅ Información de fecha seleccionada en español (FORMATO COMPLETO)
    nombre_dia_seleccionado = nombres_dias_es[fecha_seleccionada.strftime('%A')]
    nombre_mes_seleccionado = nombres_meses_es[fecha_seleccionada.month]
    
    # Formato: "Lunes 03 de Noviembre"
    fecha_formateada = f"{nombre_dia_seleccionado} {fecha_seleccionada.day:02d} de {nombre_mes_seleccionado}"

    # ✅ Filtrar funciones vigentes para la fecha seleccionada
    funciones = (
        Funcion.objects.filter(
            activa=True,
            fecha_inicio__lte=fecha_seleccionada,
        )
        .select_related('pelicula')
        .order_by('-pelicula__id', 'horario')
    )

    funciones_filtradas = []
    for funcion in funciones:
        fecha_fin_funcion = funcion.fecha_inicio + timedelta(weeks=funcion.semanas) - timedelta(days=1)
        
        # Verificar que la función esté en el rango de fechas
        if funcion.fecha_inicio <= fecha_seleccionada <= fecha_fin_funcion:
            # Si es hoy, verificar que no haya pasado (margen de 10 minutos)
            if fecha_seleccionada == hoy:
                hora_funcion = datetime.strptime(funcion.horario, '%H:%M').time()
                datetime_funcion = datetime.combine(fecha_seleccionada, hora_funcion)
                
                margen = timedelta(minutes=10)
                if datetime_funcion > (ahora_naive - margen):
                    funciones_filtradas.append(funcion)
            else:
                # Para fechas futuras, incluir todas
                funciones_filtradas.append(funcion)

    # ✅ Agrupar funciones por película (ordenadas por ID descendente)
    peliculas_cartelera = []
    peliculas_vistas = set()
    
    for funcion in funciones_filtradas:
        if funcion.pelicula.id not in peliculas_vistas:
            # Primera vez que vemos esta película
            pelicula = funcion.pelicula
            pelicula.funciones_del_dia = [f for f in funciones_filtradas if f.pelicula.id == pelicula.id]
            peliculas_cartelera.append(pelicula)
            peliculas_vistas.add(pelicula.id)

    # ✅ Enriquecer películas
    for pelicula in peliculas_cartelera:
        pelicula.generos_list = pelicula.get_generos_list()
        pelicula.salas_con_formato = pelicula.get_salas_con_formato()

    context = {
        'peliculas': peliculas_cartelera,
        'dias_disponibles': dias_disponibles,
        'fecha_seleccionada': fecha_seleccionada,
        'fecha_formateada': fecha_formateada,
        'hoy': hoy,
    }

    return render(request, 'horarios.html', context)


##################################################################
##################################################################
# Zona horaria de El Salvador
TZ_ES = pytz.timezone("America/El_Salvador")

# Distancia mínima entre funciones en minutos (2h30m)
DISTANCIA_MINIMA_MIN = 150

def _ahora_es():
    return datetime.now(TZ_ES)

def parse_hora(hhmm: str):
    """Convierte 'HH:MM' a objeto datetime.time"""
    h, m = hhmm.split(":")
    return datetime(2000, 1, 1, int(h), int(m)).time()

def _minutos_entre(hora_a, hora_b):
    a = datetime(2000, 1, 1, hora_a.hour, hora_a.minute)
    b = datetime(2000, 1, 1, hora_b.hour, hora_b.minute)
    return abs(int((b - a).total_seconds() // 60))

def hay_conflicto_distancia(hora, existentes):
    """Valida si 'hora' incumple la distancia mínima con 'existentes'"""
    for h in existentes:
        if _minutos_entre(h, hora) < DISTANCIA_MINIMA_MIN:
            return True
    return False

#********************************************
# Vista Administrar funciones
#----------------------------------------

@csrf_exempt
def administrar_funciones(request):
    hoy_es = _ahora_es().date()

    # Extraer TODAS las películas de la BD (ordenadas: última agregada primero)
    peliculas = Pelicula.objects.all().order_by('-id')

    # Enriquecer cada película con sus datos asociados
    generos_choices = dict(Pelicula.GENERO_CHOICES)
    peliculas_con_pares = []
    
    for p in peliculas:
        pares = p.get_salas_con_formato()
        generos_nombres = [generos_choices.get(g, g) for g in p.get_generos_list()]
        peliculas_con_pares.append({
            'obj': p,
            'pares': pares,
            'generos_nombres': ", ".join(generos_nombres),
            'clasificacion': p.clasificacion,
            'idioma': p.idioma
        })

    # Para el dropdown de películas
    for pelicula in peliculas:
        pelicula.generos_list = pelicula.get_generos_list()
        pelicula.salas_con_formato = pelicula.get_salas_con_formato()
        pelicula.estrellas = (
            pelicula.get_rating_estrellas()
            if hasattr(pelicula, 'get_rating_estrellas')
            else {'llenas': 0, 'media': False}
        )

    # Filtrar funciones según su fecha_fin calculada
    todas_funciones_activas = Funcion.objects.filter(
        activa=True
    ).select_related('pelicula').order_by('pelicula__id', 'horario')
    
    # Separar en vigentes y pasadas según fecha_fin
    funciones_actuales = []
    for funcion in todas_funciones_activas:
        fecha_fin = funcion.fecha_inicio + timedelta(weeks=funcion.semanas) - timedelta(days=1)
        if fecha_fin >= hoy_es:
            funciones_actuales.append(funcion)
    
    # Funciones pasadas: inactivas O con fecha_fin ya pasada
    todas_funciones = Funcion.objects.filter(
        models.Q(activa=False)
    ).select_related('pelicula').order_by('pelicula__id', 'horario')
    
    funciones_pasadas = list(todas_funciones)
    
    # Agregar funciones activas pero con fecha_fin pasada
    for funcion in todas_funciones_activas:
        fecha_fin = funcion.fecha_inicio + timedelta(weeks=funcion.semanas) - timedelta(days=1)
        if fecha_fin < hoy_es:
            funciones_pasadas.append(funcion)

    funcion_editar = None

    # ============================================================
    # FUNCIÓN AUXILIAR: Validar conflicto de horario/sala
    # ============================================================
    def hay_conflicto_horario_sala(sala_nombre, horario, fecha_inicio, semanas, funcion_id_excluir=None):
        """
        Verifica si hay conflicto de horario en la misma sala durante el período.
        Retorna True si hay conflicto, False si está libre.
        """
        fecha_fin_nueva = fecha_inicio + timedelta(weeks=int(semanas)) - timedelta(days=1)
        
        # Buscar funciones activas en la misma sala y horario
        funciones_existentes = Funcion.objects.filter(
            sala__icontains=sala_nombre,
            horario=horario,
            activa=True
        )
        
        if funcion_id_excluir:
            funciones_existentes = funciones_existentes.exclude(id=funcion_id_excluir)
        
        for func in funciones_existentes:
            fecha_fin_existente = func.fecha_inicio + timedelta(weeks=func.semanas) - timedelta(days=1)
            
            # Verificar solapamiento de períodos
            if fecha_inicio <= fecha_fin_existente and fecha_fin_nueva >= func.fecha_inicio:
                return True, func
        
        return False, None

    # ============================================================
    # PROCESAR FORMULARIOS
    # ============================================================
    if request.method == "POST":
        accion = request.POST.get("accion")

        # --- AGREGAR NUEVA FUNCIÓN CON MÚLTIPLES HORARIOS ---
        if accion == "agregar":
            pelicula_id = request.POST.get("pelicula")
            fecha_inicio_str = request.POST.get("fecha_inicio")
            semanas = request.POST.get("semanas", 1)
            
            horarios = request.POST.getlist('horario[]')
            salas = request.POST.getlist('sala[]')

            try:
                if not all([pelicula_id, fecha_inicio_str]):
                    messages.error(request, "❌ Película y fecha son obligatorios.")
                    return redirect("administrar_funciones")

                if not horarios or not salas:
                    messages.error(request, "❌ Debes agregar al menos un horario y sala.")
                    return redirect("administrar_funciones")

                pelicula = Pelicula.objects.get(id=pelicula_id)
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()

                if fecha_inicio < hoy_es:
                    messages.error(request, "❌ No puedes crear funciones con fecha pasada.")
                    return redirect("administrar_funciones")

                funciones_creadas = 0
                errores = []

                for i, (horario, sala_nombre) in enumerate(zip(horarios, salas)):
                    if not horario or not sala_nombre:
                        continue

                    try:
                        salas_pelicula = [sala_tuple[0] for sala_tuple in pelicula.get_salas_con_formato()]
                        if sala_nombre not in salas_pelicula:
                            errores.append(f"La sala '{sala_nombre}' no está disponible para esta película")
                            continue

                        # ✅ VALIDACIÓN CRÍTICA: Verificar conflicto exacto de horario/sala
                        hay_conflicto_hs, funcion_conflicto_hs = hay_conflicto_horario_sala(
                            sala_nombre, horario, fecha_inicio, semanas
                        )
                        
                        if hay_conflicto_hs:
                            fecha_fin_conflicto = funcion_conflicto_hs.fecha_inicio + timedelta(weeks=funcion_conflicto_hs.semanas) - timedelta(days=1)
                            errores.append(
                                f"❌ Ya existe '{funcion_conflicto_hs.pelicula.nombre}' en {sala_nombre} "
                                f"a las {horario} desde {funcion_conflicto_hs.fecha_inicio.strftime('%d/%m/%Y')} "
                                f"hasta {fecha_fin_conflicto.strftime('%d/%m/%Y')}"
                            )
                            continue

                        # ✅ VALIDACIÓN: Distancia mínima entre horarios (misma sala, mismo día)
                        funciones_misma_sala = Funcion.objects.filter(
                            sala__icontains=sala_nombre,
                            activa=True
                        ).exclude(horario=horario)
                        
                        fecha_fin_nueva = fecha_inicio + timedelta(weeks=int(semanas)) - timedelta(days=1)
                        
                        conflicto_distancia = False
                        for func_sala in funciones_misma_sala:
                            fecha_fin_existente = func_sala.fecha_inicio + timedelta(weeks=func_sala.semanas) - timedelta(days=1)
                            
                            # Solo validar si los períodos se solapan
                            if fecha_inicio <= fecha_fin_existente and fecha_fin_nueva >= func_sala.fecha_inicio:
                                horarios_existentes = [parse_hora(func_sala.horario)]
                                nuevo_horario = parse_hora(horario)
                                
                                if hay_conflicto_distancia(nuevo_horario, horarios_existentes):
                                    errores.append(
                                        f"⚠️ Horario {horario} en {sala_nombre}: debe haber mínimo 2h30m "
                                        f"con la función de las {func_sala.horario}"
                                    )
                                    conflicto_distancia = True
                                    break
                        
                        if conflicto_distancia:
                            continue

                        # Si no hubo conflictos, crear la función
                        nueva_funcion = Funcion(
                            pelicula=pelicula,
                            sala=sala_nombre,
                            horario=horario,
                            fecha_inicio=fecha_inicio,
                            semanas=int(semanas),
                            activa=True
                        )
                        nueva_funcion.save()
                        funciones_creadas += 1

                    except Exception as e:
                        errores.append(f"Error en horario {horario}: {str(e)}")

                if funciones_creadas > 0:
                    messages.success(request, f"✅ {funciones_creadas} función(es) agregada(s) para {pelicula.nombre}")
                
                if errores:
                    for error in errores:
                        messages.warning(request, f"⚠️ {error}")
                
                # ✅ IMPORTANTE: Siempre redirigir después de procesar POST
                return redirect("administrar_funciones")
                
            except Pelicula.DoesNotExist:
                messages.error(request, "❌ La película seleccionada no existe.")
                return redirect("administrar_funciones")
            except ValueError as e:
                messages.error(request, f"❌ Error en el formato de fecha: {str(e)}")
                return redirect("administrar_funciones")
            except Exception as e:
                messages.error(request, f"❌ Error al agregar funciones: {str(e)}")
                return redirect("administrar_funciones")

        # --- EDITAR FUNCIÓN ---
        elif accion == "editar":
            funcion_id = request.POST.get("funcion_id")
            pelicula_id = request.POST.get("pelicula")
            sala_nombre = request.POST.getlist("sala[]")[0] if request.POST.getlist("sala[]") else None
            horario = request.POST.getlist("horario[]")[0] if request.POST.getlist("horario[]") else None
            fecha_inicio_str = request.POST.get("fecha_inicio")
            semanas = request.POST.get("semanas", 1)

            try:
                funcion = Funcion.objects.get(id=funcion_id)
                
                if not pelicula_id:
                    pelicula = funcion.pelicula
                else:
                    pelicula = Pelicula.objects.get(id=pelicula_id)
                
                if not sala_nombre:
                    sala_nombre = funcion.sala
                
                if not horario:
                    horario = funcion.horario
                
                if not fecha_inicio_str:
                    fecha_inicio = funcion.fecha_inicio
                else:
                    fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()

                if fecha_inicio < hoy_es and fecha_inicio != funcion.fecha_inicio:
                    messages.error(request, "❌ No puedes programar funciones con fecha pasada.")
                    return redirect("administrar_funciones")

                salas_pelicula = [sala_tuple[0] for sala_tuple in pelicula.get_salas_con_formato()]
                if sala_nombre not in salas_pelicula:
                    messages.error(request, f"❌ La sala '{sala_nombre}' no está disponible para esta película")
                    return redirect("administrar_funciones")

                # ✅ VALIDACIÓN: Verificar conflicto de horario/sala
                hay_conflicto_hs, funcion_conflicto_hs = hay_conflicto_horario_sala(
                    sala_nombre, horario, fecha_inicio, semanas, funcion_id_excluir=funcion_id
                )
                
                if hay_conflicto_hs:
                    fecha_fin_conflicto = funcion_conflicto_hs.fecha_inicio + timedelta(weeks=funcion_conflicto_hs.semanas) - timedelta(days=1)
                    messages.error(
                        request,
                        f"❌ Ya existe '{funcion_conflicto_hs.pelicula.nombre}' en {sala_nombre} "
                        f"a las {horario} desde {funcion_conflicto_hs.fecha_inicio.strftime('%d/%m/%Y')} "
                        f"hasta {fecha_fin_conflicto.strftime('%d/%m/%Y')}"
                    )
                    return redirect("administrar_funciones")

                # ✅ VALIDACIÓN: Distancia mínima
                funciones_misma_sala = Funcion.objects.filter(
                    sala__icontains=sala_nombre,
                    activa=True
                ).exclude(id=funcion_id).exclude(horario=horario)
                
                fecha_fin_nueva = fecha_inicio + timedelta(weeks=int(semanas)) - timedelta(days=1)
                
                for func_sala in funciones_misma_sala:
                    fecha_fin_existente = func_sala.fecha_inicio + timedelta(weeks=func_sala.semanas) - timedelta(days=1)
                    
                    if fecha_inicio <= fecha_fin_existente and fecha_fin_nueva >= func_sala.fecha_inicio:
                        horarios_existentes = [parse_hora(func_sala.horario)]
                        nuevo_horario = parse_hora(horario)
                        
                        if hay_conflicto_distancia(nuevo_horario, horarios_existentes):
                            messages.error(
                                request,
                                f"❌ Horario {horario} en {sala_nombre}: debe haber mínimo 2h30m "
                                f"con la función de las {func_sala.horario}"
                            )
                            return redirect("administrar_funciones")

                # Si pasó todas las validaciones, actualizar
                funcion.pelicula = pelicula
                funcion.sala = sala_nombre
                funcion.horario = horario
                funcion.fecha_inicio = fecha_inicio
                funcion.semanas = int(semanas)
                funcion.save()

                messages.success(request, f"✏️ Función editada correctamente: {pelicula.nombre} - {sala_nombre} - {horario}")
                return redirect("administrar_funciones")
                
            except Funcion.DoesNotExist:
                messages.error(request, "❌ La función que intentas editar no existe.")
            except Pelicula.DoesNotExist:
                messages.error(request, "❌ La película seleccionada no existe.")
            except Exception as e:
                messages.error(request, f"❌ Error al editar función: {str(e)}")

        # --- ELIMINAR ---
        elif accion == "eliminar":
            funcion_id = request.POST.get("funcion_id")
            if funcion_id:
                try:
                    funcion = Funcion.objects.get(id=funcion_id)
                    nombre_pelicula = funcion.pelicula.nombre
                    
                    funcion.activa = False
                    funcion.fecha_eliminacion = hoy_es
                    funcion.save()
                    
                    messages.success(request, f"🗑️ Función de '{nombre_pelicula}' desactivada correctamente.")
                    
                except Funcion.DoesNotExist:
                    messages.error(request, "La función que intentas eliminar no existe.")
                except Exception as e:
                    messages.error(request, f"Error al eliminar la función: {str(e)}")
            else:
                messages.error(request, "No se proporcionó ID de función para eliminar.")
            return redirect("administrar_funciones")
        
        # --- ELIMINAR PERMANENTEMENTE (para funciones pasadas) ---
        elif accion == "eliminar_permanente":
            funcion_id = request.POST.get("funcion_id")
            if funcion_id:
                try:
                    funcion = Funcion.objects.get(id=funcion_id)
                    nombre_pelicula = funcion.pelicula.nombre
                    
                    # Eliminar permanentemente
                    funcion.delete()
                    
                    messages.success(request, f"🗑️ Función de '{nombre_pelicula}' eliminada permanentemente.")
                    
                except Funcion.DoesNotExist:
                    messages.error(request, "La función que intentas eliminar no existe.")
                except Exception as e:
                    messages.error(request, f"Error al eliminar la función: {str(e)}")
            else:
                messages.error(request, "No se proporcionó ID de función para eliminar.")
            return redirect("administrar_funciones")

        # --- REACTIVAR FUNCIÓN ---
        elif accion == "reactivar":
            funcion_id = request.POST.get("funcion_id")
            if funcion_id:
                try:
                    funcion = Funcion.objects.get(id=funcion_id)
                    if funcion.fecha_inicio < hoy_es:
                        funcion.fecha_inicio = hoy_es
                    
                    funcion.activa = True
                    funcion.fecha_eliminacion = None
                    funcion.save()
                    
                    messages.success(request, f"✅ Función de '{funcion.pelicula.nombre}' reactivada correctamente.")
                    
                except Funcion.DoesNotExist:
                    messages.error(request, "La función que intentas reactivar no existe.")
                except Exception as e:
                    messages.error(request, f"Error al reactivar la función: {str(e)}")
            else:
                messages.error(request, "No se proporcionó ID de función para reactivar.")
            return redirect("administrar_funciones")

    elif request.method == "GET" and "editar" in request.GET:
        funcion_id = request.GET.get("editar")
        if funcion_id:
            try:
                funcion_editar = Funcion.objects.get(id=funcion_id)
            except Funcion.DoesNotExist:
                messages.error(request, "La función que intentas editar no existe.")
        else:
            messages.error(request, "No se proporcionó ID de función para editar.")

    # Agrupar funciones por película Y fecha_inicio
    funciones_actuales_sorted = sorted(funciones_actuales, key=lambda f: (f.pelicula.id, f.fecha_inicio))
    funciones_actuales_agrupadas = []
    
    for (pelicula_id, fecha_inicio), grupo in groupby(funciones_actuales_sorted, key=lambda f: (f.pelicula.id, f.fecha_inicio)):
        funciones_lista = list(grupo)
        funciones_actuales_agrupadas.append({
            'pelicula': funciones_lista[0].pelicula,
            'fecha_inicio': fecha_inicio,
            'funciones': funciones_lista
        })

    # Para funciones pasadas
    funciones_pasadas_sorted = sorted(funciones_pasadas, key=lambda f: (f.pelicula.id, f.fecha_inicio))
    funciones_pasadas_agrupadas = []
    
    for (pelicula_id, fecha_inicio), grupo in groupby(funciones_pasadas_sorted, key=lambda f: (f.pelicula.id, f.fecha_inicio)):
        funciones_lista = list(grupo)
        funciones_pasadas_agrupadas.append({
            'pelicula': funciones_lista[0].pelicula,
            'fecha_inicio': fecha_inicio,
            'funciones': funciones_lista
        })

    return render(request, "administrar_funciones.html", {
        "peliculas": peliculas,
        "peliculas_con_pares": peliculas_con_pares,
        "funciones_actuales_agrupadas": funciones_actuales_agrupadas,
        "funciones_pasadas_agrupadas": funciones_pasadas_agrupadas,
        "funcion_editar": funcion_editar,
        "HORARIOS_DISPONIBLES": Funcion.HORARIOS_DISPONIBLES,
        "hoy_es": hoy_es,
    })



#########################################################################
### Reportes Administrativos##############################################

from django.db.models import Sum
from django.shortcuts import render
from .models import Venta, Pelicula

from django.db.models import Sum, Q
from django.shortcuts import render
from .models import Venta, Pelicula


@admin_required
def reportes_admin(request):
    peliculas = Pelicula.objects.all()
    pelicula_id = request.GET.get('pelicula')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    # ✅ USAR MODELO VENTA EN LUGAR DE RESERVA
    ventas = Venta.objects.all()

    # --- Filtros por fecha ---
    if fecha_inicio and fecha_fin:
        ventas = ventas.filter(fecha__range=[fecha_inicio, fecha_fin])

    # --- Filtro por película ---
    if pelicula_id:
        ventas = ventas.filter(pelicula_id=pelicula_id)

    # --- Agrupar por película y ordenar por total de boletos ---
    ventas_resumen = (
        ventas.values('pelicula__nombre')
        .annotate(
            total_boletos=Sum('cantidad_boletos'),
            total_venta=Sum('total_venta')
        )
        .order_by('-total_boletos')
    )

    # --- Calcular boletos por formato ---
    formatos_info = {}
    
    for v in ventas_resumen:
        pelicula_nombre = v['pelicula__nombre']
        formatos_info[pelicula_nombre] = []
        
        # Obtener reservas de esta película
        reservas_filtradas = Reserva.objects.filter(
            pelicula__nombre=pelicula_nombre,
            estado__in=['RESERVADO', 'CONFIRMADO']
        )
        
        # Aplicar mismos filtros de fecha
        if fecha_inicio and fecha_fin:
            reservas_filtradas = reservas_filtradas.filter(
                fecha_funcion__range=[fecha_inicio, fecha_fin]
            )
        
        # Agrupar por formato
        formatos_pelicula = (
            reservas_filtradas.values('formato')
            .annotate(total_boletos=Sum('cantidad_boletos'))
            .order_by('-total_boletos')
        )
        
        for f in formatos_pelicula:
            if f['formato'] and f['total_boletos'] > 0:
                formatos_info[pelicula_nombre].append({
                    'formato': f['formato'],
                    'total_boletos': f['total_boletos']
                })

    # --- Totales generales ---
    resumen_general = ventas.aggregate(
        total_boletos=Sum('cantidad_boletos'),
        total_ventas=Sum('total_venta')
    )

    # Asegurar que no sean None
    if resumen_general['total_boletos'] is None:
        resumen_general['total_boletos'] = 0
    if resumen_general['total_ventas'] is None:
        resumen_general['total_ventas'] = 0.0

    # --- Películas más populares ---
    popularidad = (
        ventas.values('pelicula__nombre')
        .annotate(total_boletos=Sum('cantidad_boletos'))
        .order_by('-total_boletos')[:5]
    )

    context = {
        'peliculas': peliculas,
        'ventas_resumen': ventas_resumen,
        'resumen_general': resumen_general,
        'popularidad': popularidad,
        'formatos_info': formatos_info,
    }

    return render(request, 'reportes_admin.html', context)


###################################################################################

### Exportar Reportes exell y pdf

def exportar_excel(request):
    # --- Obtener filtros ---
    pelicula_id = request.GET.get('pelicula')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    # --- Base de datos ---
    reservas = Reserva.objects.all()
    filtros_activos = False

    if pelicula_id:
        reservas = reservas.filter(pelicula_id=pelicula_id)
        filtros_activos = True

    if fecha_inicio and fecha_fin:
        reservas = reservas.filter(fecha_reserva__date__range=[fecha_inicio, fecha_fin])
        filtros_activos = True

    # --- Agrupar por película ---
    peliculas = reservas.values('pelicula__nombre').annotate(
        total_boletos=Sum('cantidad_boletos'),
        total_venta=Sum('precio_total')
    ).order_by('pelicula__nombre')

    # --- Calcular boletos por formato ---
    datos = []
    for peli in peliculas:
        nombre = peli['pelicula__nombre']
        total_boletos = peli['total_boletos']
        total_venta = peli['total_venta']

        # Obtener datos de formatos
        formatos = reservas.filter(pelicula__nombre=nombre).values('formato').annotate(
            total_boletos=Sum('cantidad_boletos')
        )

        # Combinar datos de formatos en una cadena
        if formatos:
            formatos_texto = ", ".join([f"{f['formato']}: {f['total_boletos']} boletos" for f in formatos])
        else:
            formatos_texto = "Sin formato"

        datos.append({
            "Película": nombre,
            "Boletos por formato": formatos_texto,
            "Total boletos": total_boletos,
            "Total ($)": float(total_venta)
        })

    # --- Calcular totales generales ---
    total_general_boletos = sum(p['Total boletos'] for p in datos)
    total_general_ventas = sum(p['Total ($)'] for p in datos)

    # Agregar fila de totales
    datos.append({
        "Película": "TOTAL GENERAL" if not filtros_activos else "TOTAL FILTRADO",
        "Boletos por formato": "",
        "Total boletos": total_general_boletos,
        "Total ($)": total_general_ventas
    })

    # --- Crear DataFrame ---
    df = pd.DataFrame(datos)

    # --- Crear respuesta con archivo Excel ---
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Nombre dinámico del archivo
    if filtros_activos:
        response["Content-Disposition"] = 'attachment; filename="reporte_filtrado_cinedot.xlsx"'
        hoja_nombre = "Reporte Filtrado"
    else:
        response["Content-Disposition"] = 'attachment; filename="reporte_general_cinedot.xlsx"'
        hoja_nombre = "Reporte General"

    # --- Exportar a Excel ---
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=hoja_nombre)

        hoja = writer.sheets[hoja_nombre]

        # Ajustar ancho de columnas
        for columna in hoja.columns:
            hoja.column_dimensions[columna[0].column_letter].width = 30

        # Fijar encabezados
        hoja.freeze_panes = "A2"

        # Estilo más profesional
        from openpyxl.styles import Font, PatternFill, Alignment

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        for cell in hoja[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Centrar contenido
        for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row):
            for celda in fila:
                celda.alignment = Alignment(horizontal="center", vertical="center")

    return response


### Reportes PDF#########

def exportar_pdf(request):
    # --- Obtener filtros ---
    pelicula_id = request.GET.get('pelicula')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    # --- Base de datos ---
    reservas = Reserva.objects.all()
    filtros_activos = False

    if pelicula_id:
        reservas = reservas.filter(pelicula_id=pelicula_id)
        filtros_activos = True

    if fecha_inicio and fecha_fin:
        reservas = reservas.filter(fecha_reserva__date__range=[fecha_inicio, fecha_fin])
        filtros_activos = True

    # --- Crear respuesta PDF ---
    response = HttpResponse(content_type='application/pdf')

    if filtros_activos:
        response['Content-Disposition'] = 'attachment; filename="reporte_filtrado.pdf"'
        titulo = "Reporte Filtrado de Ventas"
    else:
        response['Content-Disposition'] = 'attachment; filename="reporte_general.pdf"'
        titulo = "Reporte General de Ventas (Todas las Películas)"

    # --- Crear PDF en memoria ---
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # --- Encabezado ---
    p.setFont("Helvetica-Bold", 16)
    p.drawString(130, height - 50, f"🎬 {titulo}")

    p.setFont("Helvetica", 10)
    p.drawString(50, height - 70, "Generado automáticamente por el sistema Cinedot")
    p.drawString(50, height - 85, f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    y = height - 110

    # --- Mostrar criterios de filtro ---
    if filtros_activos:
        p.setFont("Helvetica-Oblique", 9)
        if pelicula_id:
            peli = Pelicula.objects.get(id=pelicula_id)
            p.drawString(50, y, f"Película: {peli.nombre}")
            y -= 12
        if fecha_inicio and fecha_fin:
            p.drawString(50, y, f"Rango de fechas: {fecha_inicio} a {fecha_fin}")
            y -= 12
        y -= 10

    # --- Encabezado de tabla ---
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y, "Película")
    p.drawString(220, y, "Boletos por formato")
    p.drawString(420, y, "Total boletos")
    p.drawString(510, y, "Total ($)")
    y -= 15
    p.line(50, y, 560, y)
    y -= 20

    # --- Cálculo de datos por película ---
    total_general_boletos = 0
    total_general_ventas = 0

    if reservas.exists():
        # Agrupar por película
        peliculas = reservas.values('pelicula__nombre').annotate(
            total_boletos=Sum('cantidad_boletos'),
            total_venta=Sum('precio_total')
        ).order_by('pelicula__nombre')

        for peli in peliculas:
            nombre = peli['pelicula__nombre']

            # Formatos disponibles por película
            formatos = reservas.filter(pelicula__nombre=nombre).values('formato').annotate(
                total_boletos=Sum('cantidad_boletos')
            )

            # Calcular total
            total_boletos = peli['total_boletos']
            total_venta = peli['total_venta']

            # Mostrar nombre
            p.setFont("Helvetica-Bold", 10)
            p.drawString(50, y, nombre)
            y -= 15

            # Mostrar formatos
            p.setFont("Helvetica", 9)
            if formatos:
                for f in formatos:
                    texto = f"{f['formato']}: {f['total_boletos']} boletos"
                    p.drawString(70, y, texto)
                    y -= 12
            else:
                p.drawString(70, y, "Sin formato")
                y -= 12

            # Mostrar totales por película
            p.setFont("Helvetica-Bold", 10)
            p.drawString(420, y + 5, str(total_boletos))
            p.drawString(510, y + 5, f"${total_venta:.2f}")
            y -= 18

            # Línea divisoria
            p.setFont("Helvetica", 9)
            p.line(50, y, 560, y)
            y -= 10

            total_general_boletos += total_boletos
            total_general_ventas += float(total_venta)

            # Salto de página si se llena
            if y < 100:
                p.showPage()
                y = height - 100

    else:
        p.drawString(50, y, "No hay datos disponibles con los filtros seleccionados.")
        y -= 20

    # --- Totales generales ---
    p.setFont("Helvetica-Bold", 11)
    y -= 10
    p.line(50, y, 560, y)
    y -= 20
    p.drawString(50, y, "Totales generales:")
    p.drawString(420, y, str(total_general_boletos))
    p.drawString(510, y, f"${total_general_ventas:.2f}")

    # --- Guardar PDF ---
    p.showPage()
    p.save()

    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

from django.contrib.auth.views import PasswordResetView
from django.contrib.auth.forms import PasswordResetForm
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.contrib.sites.shortcuts import get_current_site
from django.urls import reverse
from urllib.parse import urlencode
from myapp.email import send_brevo_email  # la función que creamos


class CustomPasswordResetView(PasswordResetView):
    email_template_name = 'registration/password_reset_email.html'  # plantilla del correo
    subject_template_name = 'registration/password_reset_subject.txt'  # asunto

    def send_mail(self, subject, email_template_name, context, from_email, to_email, html_email_template_name=None):
        # Renderizamos el contenido del correo
        html_content = render_to_string(email_template_name, context)
        send_brevo_email(
            to_emails=[to_email],
            subject=subject,
            html_content=html_content
        )



import json

##########################################################################################
@admin_required
def dashboard_admin(request):
    # ✅ USAR MODELO VENTA
    ventas = Venta.objects.all()
    
    # Top 10 películas más vendidas
    top_peliculas = list(
        ventas.values('pelicula__nombre')
        .annotate(
            total_boletos=Sum('cantidad_boletos'),
            total_venta=Sum('total_venta')
        )
        .order_by('-total_boletos')[:10]
    )

    # Para formatos, usar Reservas
    reservas_confirmadas = Reserva.objects.filter(
        estado__in=['RESERVADO', 'CONFIRMADO']
    )
    
    formatos = list(
        reservas_confirmadas.values('formato')
        .annotate(
            total_boletos=Sum('cantidad_boletos'),
            total_venta=Sum('precio_total')
        )
        .order_by('-total_boletos')
    )

    # Filtrar formatos vacíos
    formatos = [f for f in formatos if f['formato']]

    # Resumen general
    resumen_general = ventas.aggregate(
        total_boletos=Sum('cantidad_boletos'),
        total_ventas=Sum('total_venta')
    )

    # Valores por defecto
    if resumen_general['total_boletos'] is None:
        resumen_general['total_boletos'] = 0
    if resumen_general['total_ventas'] is None:
        resumen_general['total_ventas'] = 0.0

    # Debug en consola
    print("=" * 50)
    print("📊 DASHBOARD DATA:")
    print(f"Total Boletos: {resumen_general['total_boletos']}")
    print(f"Total Ventas: ${resumen_general['total_ventas']:.2f}")
    print(f"Top Películas: {len(top_peliculas)}")
    print(f"Formatos: {len(formatos)}")
    print("=" * 50)

    context = {
        "resumen_general": resumen_general,
        "top_peliculas": top_peliculas,
        "formatos": formatos,
        "top_peliculas_json": json.dumps(top_peliculas, default=str),
        "formatos_json": json.dumps(formatos, default=str),
    }
    
    return render(request, "dashboard_admin.html", context)


##############################################################################

@login_required
def mis_reservaciones_cancelables(request):
    
    TIEMPO_MAXIMO_CANCELACION = timezone.now() - timedelta(hours=24)
    HORA_LIMITE_CANCELACION = timezone.now() + timedelta(hours=3)
    # Filtra las reservas solo de usuarios logueados
    reservas_posibles = Reserva.objects.filter(
        usuario=request.user,
        fecha_reserva__gte=TIEMPO_MAXIMO_CANCELACION, 
        estado__in=['RESERVADO', 'CONFIRMADO'],
    ).order_by('-fecha_reserva')
    reservas_validas = []
    
    for reserva in reservas_posibles:
       
        try:
            
            hora_funcion = datetime.strptime(reserva.horario, '%H:%M').time()
            datetime_funcion = datetime.combine(reserva.fecha_funcion, hora_funcion)
            
            if timezone.is_aware(timezone.now()):
                datetime_funcion = timezone.make_aware(datetime_funcion)
            
            # Comprobación de las 3 horas: La función debe ser en MÁS de 3 horas
            if datetime_funcion > HORA_LIMITE_CANCELACION:
                # Formato AM/PM: 
                horario_ampm = datetime_funcion.strftime('%I:%M %p') 
            
            # Adjuntar el campo combinado para la tabla
                sala_limpia = str(reserva.sala).replace('Sala', '').strip()
            
                reserva.horario_sala_web = f"{horario_ampm} / Sala {sala_limpia}"
                reservas_validas.append(reserva)
        
        except ValueError:
            # Si el formato del horario no es correcto, simplemente omite la reserva
            continue
    
    context = {
        'reservas': reservas_validas,
        'tiempo_limite_cancelacion': 24
    }
    return render(request, 'reservaciones_cancelables.html', context)

@login_required
def cancelar_reserva(request, pk):
    reserva = get_object_or_404(Reserva, pk=pk, usuario=request.user)
    
    # Constantes de los límites
    TIEMPO_MAXIMO_CANCELACION = timedelta(hours=24) # 24h desde la compra
    TIEMPO_MINIMO_CANCELACION_PROYECCION = timedelta(hours=3) # 3h antes de la función
    
    tiempo_transcurrido = timezone.now() - reserva.fecha_reserva
    
    try:
        hora_funcion = datetime.strptime(reserva.horario, '%H:%M').time()
        datetime_funcion = datetime.combine(reserva.fecha_funcion, hora_funcion)
        
        if timezone.is_aware(timezone.now()):
            datetime_funcion = timezone.make_aware(datetime_funcion)

        tiempo_hasta_proyeccion = datetime_funcion - timezone.now()
        fecha_formateada = reserva.fecha_funcion.strftime('%d/%m/%Y')
        
        
        horario_ampm = datetime_funcion.strftime('%I:%M %p')
        
        
        horario_sala_combinado = f"{horario_ampm} "
    except ValueError:
        messages.error(request, 'Error interno: No se pudo verificar la hora de la función.')
        return redirect('mis_reservaciones_cancelables')
    
    
    if request.method == 'POST':
        
        #  VERIFICACIÓN DE LA REGLA DE 24 HORAS DESDE LA COMPRA
        if tiempo_transcurrido > TIEMPO_MAXIMO_CANCELACION:
            messages.error(request, 'Esta reserva ya no puede ser cancelada: han pasado más de 24 horas desde la compra.')
            return redirect('mis_reservaciones_cancelables')
        
        # 2VERIFICACIÓN DE LA REGLA DE 3 HORAS ANTES DE LA FUNCIÓN
        # Si la reserva fue hecha dentro de las 24 horas, aún debe cumplir la regla de la función.
        if tiempo_hasta_proyeccion < TIEMPO_MINIMO_CANCELACION_PROYECCION:
            messages.error(request, 'Esta reserva ya no puede ser cancelada: la función inicia en menos de 3 horas.')
            return redirect('mis_reservaciones_cancelables')
            
        # 3. CANCELACIÓN (Si pasó ambas verificaciones)
        reserva.estado = 'CANCELADO'
        reserva.usado = True
        reserva.save()

        try:
            subject = f"Cancelación Exitosa - Reserva {reserva.codigo_reserva}"
            body = (
                f"Hola {reserva.nombre_cliente} {reserva.apellido_cliente},<br><br>"
        
                f"Confirmamos que tu reserva para la película '<strong>{reserva.pelicula.nombre}</strong>' ha sido **CANCELADA exitosamente**.<br><br>"
        
                f"<strong>Detalles de la reserva cancelada:</strong><br>"
                f"🎫 Código de reserva: <strong>{reserva.codigo_reserva}</strong><br>"
                f"🎬 Película: {reserva.pelicula.nombre}<br>"
                f"🗓️ <strong>Fecha de Función:</strong> {reserva.fecha_funcion}<br>"
                f"🕐 Horario: {horario_sala_combinado}<br>"
                f"🎬 Sala: {reserva.sala}<br>"
                f"💺 Asientos Liberados: {reserva.asientos}<br><br>"
        
                f"<strong>Información de Reembolso:</strong><br>"
                f"💸 Monto a reembolsar: <strong>${reserva.precio_total:.2f}</strong><br>"
                f"El proceso de reembolso se iniciará en las próximas 48 horas hábiles (el tiempo puede variar según tu banco).<br><br>"
        
                f"Lamentamos que no puedas asistir esta vez.<br>"
                f"¡Esperamos verte pronto en CineDot! 🍿"
            )
            
            
            send_brevo_email(
                to_emails=[reserva.email],
                subject=subject,
                html_content=body.replace("\n", "<br>"),
            )
            
        except Exception as e:
            
            print(f" Error al enviar correo de cancelación: {e}") 
            messages.warning(request, f'Reserva cancelada, pero hubo un problema al enviar la notificación por correo.')

        messages.success(request, f'La reserva {reserva.codigo_reserva} ha sido cancelada con éxito.')
        return redirect('mis_reservaciones_cancelables')
        
        return redirect('mis_reservaciones_cancelables')
        
        #messages.success(request, f'La reserva {reserva.codigo_reserva} ha sido cancelada con éxito.')
        #return redirect('mis_reservaciones_cancelables')
        
    # Si es GET, simplemente se redirige o se pide confirmación
   # return redirect('mis_reservaciones_cancelables')


#################################################################
# GESTIÓN DE MÉTODOS DE PAGO (PBI-27)
#################################################################

@login_required
def mis_metodos_pago(request):
    """
    Vista para listar los métodos de pago guardados del usuario
    """
    metodos_pago = MetodoPago.objects.filter(
        usuario=request.user,
        activo=True
    ).order_by('-es_predeterminado', '-fecha_creacion')
    
    context = {
        'metodos_pago': metodos_pago
    }
    
    return render(request, 'mis_metodos_pago.html', context)


@login_required
def agregar_metodo_pago(request):
    """
    Vista para agregar un nuevo método de pago
    """
    from datetime import datetime
    
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        alias = request.POST.get('alias', '').strip()
        es_predeterminado = request.POST.get('es_predeterminado') == 'true'
        
        errores = []
        
        # Validaciones comunes
        if not alias:
            errores.append("El alias es obligatorio")
        
        # Verificar que el alias no esté duplicado para este usuario
        if MetodoPago.objects.filter(usuario=request.user, alias=alias, activo=True).exists():
            errores.append(f"Ya tienes un método con el alias '{alias}'")
        
        if tipo == 'TARJETA':
            numero_tarjeta = request.POST.get('numero_tarjeta', '').replace(' ', '')
            mes_expiracion = request.POST.get('mes_expiracion')
            anio_expiracion = request.POST.get('anio_expiracion')
            nombre_titular = request.POST.get('nombre_titular', '').strip()
            
            # Validaciones de tarjeta
            if not numero_tarjeta or len(numero_tarjeta) < 13:
                errores.append("Número de tarjeta inválido")
            if not mes_expiracion or not anio_expiracion:
                errores.append("Fecha de expiración incompleta")
            if not nombre_titular:
                errores.append("El nombre del titular es obligatorio")
            
            # Validar que la tarjeta no esté expirada
            if mes_expiracion and anio_expiracion:
                hoy = datetime.now()
                if int(anio_expiracion) < hoy.year or \
                   (int(anio_expiracion) == hoy.year and int(mes_expiracion) < hoy.month):
                    errores.append("La tarjeta está expirada")
            
            if not errores:
                # Procesar datos de tarjeta
                tipo_tarjeta = get_card_type(numero_tarjeta)
                
                # Encriptar datos completos de la tarjeta
                fecha_expiracion = f"{int(mes_expiracion):02d}/{str(anio_expiracion)[-2:]}"
                datos_encriptados = encrypt_card_data_full(
                    numero_tarjeta=numero_tarjeta,
                    nombre_titular=nombre_titular,
                    fecha_expiracion=fecha_expiracion
                )
                
                # Crear método de pago
                metodo = MetodoPago(
                    usuario=request.user,
                    tipo='TARJETA',
                    alias=alias,
                    es_predeterminado=es_predeterminado,
                    ultimos_4_digitos=numero_tarjeta[-4:],
                    tipo_tarjeta=tipo_tarjeta,
                    mes_expiracion=int(mes_expiracion),
                    anio_expiracion=int(anio_expiracion),
                    nombre_titular=nombre_titular.upper(),
                    datos_encriptados=datos_encriptados
                )
                metodo.save()
                
                messages.success(request, f"Método de pago '{alias}' guardado exitosamente")
                return redirect('mis_metodos_pago')
                
        elif tipo == 'CUENTA_DIGITAL':
            tipo_cuenta = request.POST.get('tipo_cuenta')
            email_cuenta = request.POST.get('email_cuenta', '').strip()
            
            # Validaciones de cuenta digital
            if not tipo_cuenta:
                errores.append("Selecciona el tipo de cuenta")
            if not email_cuenta or '@' not in email_cuenta:
                errores.append("Email de cuenta inválido")
            
            if not errores:
                # Crear método de pago
                metodo = MetodoPago(
                    usuario=request.user,
                    tipo='CUENTA_DIGITAL',
                    alias=alias,
                    es_predeterminado=es_predeterminado,
                    tipo_cuenta=tipo_cuenta,
                    email_cuenta=email_cuenta
                )
                metodo.save()
                
                messages.success(request, f"Método de pago '{alias}' guardado exitosamente")
                return redirect('mis_metodos_pago')
        
        # Si hay errores, mostrarlos
        for error in errores:
            messages.error(request, error)
    
    # Contexto para el template
    context = {
        'current_year': datetime.now().year
    }
    
    return render(request, 'agregar_metodo_pago.html', context)


@login_required
def editar_metodo_pago(request, metodo_id):
    """
    Vista para editar un método de pago existente
    """
    from datetime import datetime
    
    # Verificar que el método pertenece al usuario
    metodo = get_object_or_404(MetodoPago, id=metodo_id, usuario=request.user, activo=True)
    
    if request.method == 'POST':
        alias = request.POST.get('alias', '').strip()
        es_predeterminado = request.POST.get('es_predeterminado') == 'true'
        
        errores = []
        
        # Validar alias
        if not alias:
            errores.append("El alias es obligatorio")
        
        # Verificar que el alias no esté duplicado (excepto el actual)
        if MetodoPago.objects.filter(
            usuario=request.user, 
            alias=alias, 
            activo=True
        ).exclude(id=metodo_id).exists():
            errores.append(f"Ya tienes otro método con el alias '{alias}'")
        
        if metodo.tipo == 'TARJETA':
            mes_expiracion = request.POST.get('mes_expiracion')
            anio_expiracion = request.POST.get('anio_expiracion')
            nombre_titular = request.POST.get('nombre_titular', '').strip()
            
            if not mes_expiracion or not anio_expiracion:
                errores.append("Fecha de expiración incompleta")
            if not nombre_titular:
                errores.append("El nombre del titular es obligatorio")
            
            # Validar que la tarjeta no esté expirada
            if mes_expiracion and anio_expiracion:
                hoy = datetime.now()
                if int(anio_expiracion) < hoy.year or \
                   (int(anio_expiracion) == hoy.year and int(mes_expiracion) < hoy.month):
                    errores.append("La tarjeta está expirada")
            
            if not errores:
                metodo.alias = alias
                metodo.es_predeterminado = es_predeterminado
                metodo.mes_expiracion = int(mes_expiracion)
                metodo.anio_expiracion = int(anio_expiracion)
                metodo.nombre_titular = nombre_titular.upper()
                metodo.save()
                
                messages.success(request, f"Método '{alias}' actualizado exitosamente")
                return redirect('mis_metodos_pago')
                
        elif metodo.tipo == 'CUENTA_DIGITAL':
            email_cuenta = request.POST.get('email_cuenta', '').strip()
            
            if not email_cuenta or '@' not in email_cuenta:
                errores.append("Email de cuenta inválido")
            
            if not errores:
                metodo.alias = alias
                metodo.es_predeterminado = es_predeterminado
                metodo.email_cuenta = email_cuenta
                metodo.save()
                
                messages.success(request, f"Método '{alias}' actualizado exitosamente")
                return redirect('mis_metodos_pago')
        
        # Mostrar errores
        for error in errores:
            messages.error(request, error)
    
    # Contexto para el template
    context = {
        'metodo': metodo,
        'current_year': datetime.now().year
    }
    
    return render(request, 'editar_metodo_pago.html', context)


@login_required
def eliminar_metodo_pago(request, metodo_id):
    """
    Vista para eliminar (soft delete) un método de pago
    """
    # Verificar que el método pertenece al usuario
    metodo = get_object_or_404(MetodoPago, id=metodo_id, usuario=request.user, activo=True)
    
    if request.method == 'POST':
        alias_eliminado = metodo.alias
        era_predeterminado = metodo.es_predeterminado
        
        # Soft delete: marcar como inactivo
        metodo.activo = False
        metodo.save()
        
        # Si era predeterminado, marcar otro como predeterminado
        if era_predeterminado:
            siguiente_metodo = MetodoPago.objects.filter(
                usuario=request.user,
                activo=True
            ).order_by('-fecha_creacion').first()
            
            if siguiente_metodo:
                siguiente_metodo.es_predeterminado = True
                siguiente_metodo.save()
        
        messages.success(request, f"Método '{alias_eliminado}' eliminado exitosamente")
        return redirect('mis_metodos_pago')
    
    # Si es GET, redirigir a la lista
    return redirect('mis_metodos_pago')


@login_required
def marcar_predeterminado(request, metodo_id):
    """
    Vista para marcar un método como predeterminado
    """
    # Verificar que el método pertenece al usuario
    metodo = get_object_or_404(MetodoPago, id=metodo_id, usuario=request.user, activo=True)
    
    if request.method == 'POST':
        # Desmarcar todos los métodos del usuario
        MetodoPago.objects.filter(
            usuario=request.user,
            activo=True
        ).update(es_predeterminado=False)
        
        # Marcar el seleccionado
        metodo.es_predeterminado = True
        metodo.save()
        
        messages.success(request, f"'{metodo.alias}' marcado como predeterminado")
    
    return redirect('mis_metodos_pago')


# ═══════════════════════════════════════════════════════════════
# PBI-29: GESTIÓN DE USUARIOS
# ═══════════════════════════════════════════════════════════════

@login_required
@user_passes_test(lambda u: u.is_staff)
def administrar_usuarios(request):
    """
    Vista para administrar usuarios del sistema (solo para staff)
    """
    usuarios = User.objects.all().order_by('-is_staff', 'username')
    usuario_editar = None

    # --- CREAR / EDITAR / ELIMINAR ---
    if request.method == "POST":
        accion = request.POST.get("accion")

        # Crear usuario
        if accion == "crear":
            username = request.POST.get("username", "").strip()
            email = request.POST.get("email", "").strip()
            password = request.POST.get("password", "").strip()
            es_admin = request.POST.get("es_admin") == "on"

            # Validación: evitar nombres duplicados sin importar mayúsculas/minúsculas
            if User.objects.filter(username__iexact=username).exists():
                messages.error(request, f"⚠️ El nombre de usuario '{username}' ya está registrado (sin importar mayúsculas).")
                return redirect("administrar_usuarios")

            # Convertir explícitamente el string a booleano
            is_staff = request.POST.get("is_staff") == "True"

            if not (username and email and password):
                messages.error(request, "⚠️ Todos los campos son obligatorios.")
                return redirect("administrar_usuarios")

            if User.objects.filter(username=username).exists():
                messages.warning(request, "⚠️ El nombre de usuario ya existe.")
                return redirect("administrar_usuarios")

            # Por defecto el usuario recién creado estará activo
            new_user = User.objects.create_user(
                username=username, email=email, password=password, is_staff=is_staff
            )
            new_user.is_active = True
            new_user.save()

            messages.success(request, "✅ Usuario creado correctamente.")
            return redirect("administrar_usuarios")

        # Editar usuario
        elif accion == "editar":
            user_id = request.POST.get("user_id")
            user = get_object_or_404(User, id=user_id)

            user.username = request.POST.get("username", user.username)
            user.email = request.POST.get("email", user.email)

            # Convertir explícitamente los flags a booleanos
            user.is_staff = request.POST.get("is_staff") == "True"
            # Si el select no se envía por alguna razón, conservar el valor actual
            is_active_post = request.POST.get("is_active")
            if is_active_post is not None:
                user.is_active = is_active_post == "True"

            user.save()
            messages.success(request, "✏️ Usuario actualizado correctamente.")
            return redirect("administrar_usuarios")

        # Eliminar usuario
        elif accion == "eliminar":
            user_id = request.POST.get("user_id")
            user = get_object_or_404(User, id=user_id)
            user.delete()
            messages.success(request, "🗑️ Usuario eliminado correctamente.")
            return redirect("administrar_usuarios")
        
        # Restablecer contraseña
        elif accion == "reset_password":
            user_id = request.POST.get("user_id")
            nueva_password = request.POST.get("nueva_password", "").strip()
            
            if not user_id or not nueva_password:
                messages.error(request, "⚠️ Datos incompletos para restablecer contraseña.")
                return redirect("administrar_usuarios")
            
            if len(nueva_password) < 4:
                messages.error(request, "⚠️ La contraseña debe tener al menos 4 caracteres.")
                return redirect("administrar_usuarios")
            
            try:
                user = get_object_or_404(User, id=user_id)
                user.set_password(nueva_password)
                user.save()
                messages.success(request, f"✅ Contraseña de '{user.username}' restablecida correctamente.")
            except Exception as e:
                messages.error(request, f"❌ Error al restablecer contraseña: {str(e)}")
                return redirect("administrar_usuarios")

    # --- MODO EDICIÓN ---
    if request.method == "GET" and "editar" in request.GET:
        usuario_id = request.GET.get("editar")
        usuario_editar = get_object_or_404(User, id=usuario_id)

    return render(request, "administrar_usuarios.html", {
        "usuarios": usuarios,
        "usuario_editar": usuario_editar,
    })

@staff_member_required
def estadisticas_peliculas(request):
    data = _obtener_datos_filtrados(request)
    
    context = {
        # Los datos ya están filtrados por fecha y película 
        'datos_estadisticos': data['datos_estadisticos'],
        'total_usuarios': data['total_usuarios'],
        'peliculas_list': data['peliculas_list'], 
    }

    return render(request, 'estadisticas_peliculas.html', context)

def _obtener_datos_filtrados(request):
    
    # Obtener filtros de la URL 
    pelicula_id = request.GET.get('pelicula')
    fecha_desde_str = request.GET.get('desde')
    fecha_hasta_str = request.GET.get('hasta')
    
    # Obtener datos base 
    reservas_base = Reserva.objects.filter(estado__in=['CONFIRMADO', 'RESERVADO'])
    
    # filtro de Película
    if pelicula_id:
        reservas_base = reservas_base.filter(pelicula_id=pelicula_id)
        
    #  filtro de Fechas 
    if fecha_desde_str:
        try:
            fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
            reservas_base = reservas_base.filter(fecha_reserva__gte=fecha_desde) 
        except ValueError:
            pass
            
    if fecha_hasta_str:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
            reservas_base = reservas_base.filter(fecha_reserva__lte=fecha_hasta)
        except ValueError:
            pass

    # Logica de estadistica
    
    # Obtener todas las películas para las que hubo alguna reserva en el período/filtro
    peliculas_filtradas_ids = reservas_base.values_list('pelicula', flat=True).distinct()
    
    # Calcular estadísticas por película
    datos_estadisticos = []
    total_usuarios_registrados = User.objects.filter(is_active=True).count()
    
    for pelicula_id in peliculas_filtradas_ids:
        pelicula = Pelicula.objects.get(pk=pelicula_id)
        reservas_pelicula = reservas_base.filter(pelicula=pelicula)
        
        # Calcula Boletos Vendidos
        boletos_vendidos = reservas_pelicula.aggregate(total_boletos=Sum('cantidad_boletos'))['total_boletos'] or 0
        
        # Calcula Usuarios Únicos
        usuarios_unicos = reservas_pelicula.values('usuario').distinct().count()
        
        # Calcula Puntuación Promedio
        puntuacion_obj = Valoracion.objects.filter(pelicula=pelicula).aggregate(promedio=Avg('rating'))
        puntuacion_promedio = f"{puntuacion_obj['promedio']:.2f}" if puntuacion_obj['promedio'] else 'N/A'
        
        # Calcula % Compradores
        porcentaje_compradores = 0.0
        if total_usuarios_registrados > 0:
            porcentaje_compradores = (usuarios_unicos / total_usuarios_registrados) * 100
        
        datos_estadisticos.append({
            'nombre': pelicula.nombre,
            'puntuacion_promedio': puntuacion_promedio,
            'numero_reservas': boletos_vendidos,
            'usuarios_unicos': usuarios_unicos,
            'porcentaje_compradores': f"{porcentaje_compradores:.2f}",
        })
        
    return {
        'datos_estadisticos': datos_estadisticos,
        'total_usuarios': total_usuarios_registrados,
        'peliculas_list': Pelicula.objects.all().order_by('nombre')
    }

def exportar_pdf_peliculas(request):
    data = _obtener_datos_filtrados(request)
    datos_estadisticos = data['datos_estadisticos']
    
    # Crea un buffer en memoria
    buffer = io.BytesIO()
    
    # Crea el objeto Canvas
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Encabezado del documento
    p.setFont("Helvetica-Bold", 16)
    p.setFillColor(colors.HexColor('#6a1b9a'))
    p.drawString(30, height - 30, " Reporte de Desempeño de Películas")
    
    p.setFont("Helvetica", 10)
    p.setFillColor(colors.black)
    p.drawString(30, height - 50, f"Generado: {date.today().strftime('%d/%m/%Y')}")
    p.drawString(30, height - 65, f"Total de Usuarios Registrados: {data['total_usuarios']}")
    
    # Prepara los  datos para la tabla PDF
    table_data = [
        ['Película', 'Puntuación Promedio', 'Boletos Vendidos', 'Usuarios Únicos', '% Compradores']
    ]
    for dato in datos_estadisticos:
        table_data.append([
            dato['nombre'],
            dato['puntuacion_promedio'],
            str(dato['numero_reservas']),
            str(dato['usuarios_unicos']),
            f"{dato['porcentaje_compradores']}%"
        ])

    # Crea la Tabla
    table = Table(table_data, colWidths=[150, 120, 90, 80, 80]) 
    
    # Estilo de la tabla 
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)

    # Dibuja la tabla
    table_height = len(table_data) * 20 + 30 # Altura aproximada
    table.wrapOn(p, width, table_height)
    table.drawOn(p, 30, height - 100 - table_height)

    # Finaliza el PDF y envia
    p.showPage()
    p.save()
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_peliculas_{date.today().isoformat()}.pdf"'
    return response

def exportar_excel_peliculas(request):
    #Obteniendo los datos filtrados 
    data = _obtener_datos_filtrados(request)
    datos_estadisticos = data['datos_estadisticos']
    
    # creacion de libro en excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Desempeño de Películas"

    # 3. Encabezados del excel
    headers = ['Película', 'Puntuación Promedio', 'Boletos Vendidos', 'Usuarios Únicos', '% Compradores']
    ws.append(headers)

    # 4. Datos del reporte (A partir de la fila 2)
    for dato in datos_estadisticos:
        # Aseguramos que los números se guarden como números (float o int)
        try:
            # Convierte 'N/A' a None para que Excel lo maneje como una celda vacía o de error
            puntuacion = float(dato['puntuacion_promedio']) if dato['puntuacion_promedio'] != 'N/A' else None
        except ValueError:
            puntuacion = None

        try:
            
            porcentaje = float(dato['porcentaje_compradores']) / 100.0
        except ValueError:
            porcentaje = 0.0
            
        ws.append([
            dato['nombre'],
            puntuacion,
            int(dato['numero_reservas']),
            int(dato['usuarios_unicos']),
            porcentaje
        ])


    # Definir Estilos
    # Usamos un color similar a tu encabezado de tabla (gris oscuro)
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid") 
    header_font = Font(bold=True, color="FFFFFF") 
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True) # Alineación central
    left_align = Alignment(horizontal="left", vertical="center")

    # Aplicar estilos a Encabezados (Fila 1)
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Aplicar formatos de celda (porcentaje, decimales) y alineación
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        
        porcentaje_cell = row[4] 
        # Formato de porcentaje a dos decimales
        porcentaje_cell.number_format = '0.00%' 
        
        row[1].number_format = '0.00' 
        
        # Aplicar alineación
        row[0].alignment = left_align 
        row[1].alignment = center_align 
        row[2].alignment = center_align 
        row[3].alignment = center_align 
        porcentaje_cell.alignment = center_align 


    # Ajustar ancho de columnas para mejor visualización
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                # Comprobar la longitud del valor de la celda
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Aumentar el ancho de la columna 
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width


    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Crea respuesta HTTP para la descarga
    response = HttpResponse(
        buffer.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Nombre del archivo
    response['Content-Disposition'] = f'attachment; filename="reporte_peliculas_{date.today().isoformat()}.xlsx"'
    
    return response
    